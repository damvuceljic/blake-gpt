from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from typing import Any
from pathlib import Path

from finance_copilot.common import read_json, utc_now_iso, write_json

DEFAULT_HOTQ_SCORING_CONFIG: dict[str, Any] = {
    "weights": {
        "pnl_delivery": 30,
        "variance_explainability": 25,
        "forecast_reliability": 20,
        "narrative_integrity": 15,
        "data_confidence": 10,
    },
    "thresholds": {"green_min": 80, "yellow_min": 65},
    "materiality": {"currency_mm": 0.5, "percent": 1.0},
    "penalties": {
        "lineage_degraded": 10,
        "notes_missing_ratio_high": 8,
        "missing_pack_components": 25,
    },
}

DEFAULT_HOTQ_POLICY: dict[str, Any] = {
    "policy_version": "2026.02.phase1",
    "term_guard": {
        "allowlist": ["data quality", "quality check"],
        "rewrite_mode": "deterministic",
        "banned_terms": [
            {
                "id": "quality_family",
                "pattern": r"\b(?:earnings|margin|le(?:\s+base)?)\s+quality\b",
                "rewrite": "documented variance drivers",
            },
            {
                "id": "quality_shorthand",
                "pattern": r"\bquality\b",
                "rewrite": "variance composition",
            },
            {
                "id": "core_operating_demand",
                "pattern": r"\bcore operating demand\b",
                "rewrite": "documented traffic and ticket drivers",
            },
            {
                "id": "elasticity_risk",
                "pattern": r"\belasticity risk\b",
                "rewrite": "price-volume response risk",
                "allow_if_numeric": True,
            },
            {
                "id": "traffic_sales_conversion",
                "pattern": r"\btraffic\s*/\s*sales\s*conversion\b",
                "rewrite": "traffic-to-sales bridge",
            },
            {
                "id": "trust_attack",
                "pattern": r"why should leadership trust this le base now\??",
                "rewrite": (
                    "what assumptions changed versus prior LE, who owns them, "
                    "and what lock controls apply next cycle?"
                ),
            },
        ],
    },
    "scope_rules": {
        "restaurant_first_required": True,
        "exclude_cpg_primary_without_bridge": True,
        "restaurant_tokens": [
            "traffic",
            "ticket",
            "mix",
            "pricing",
            "labor",
            "restaurant",
            "franchise",
            "sales",
            "aoi",
            "ebitda",
            "sst",
            "sss",
        ],
        "cpg_tokens": [
            "cpg",
            "consumer packaged",
            "supply chain",
            "distribution center",
            "co-manufacturing",
            "wholesale",
        ],
        "quantified_bridge_regex": r"\$?\d[\d,]*(?:\.\d+)?%?|\b\d+(?:\.\d+)?\s*(?:mm|pp)\b",
    },
    "semantics": {
        "preview_equals_le": True,
        "primary_bases": ["vs Budget", "vs LE"],
        "le_change_label": "vs prior LE",
    },
    "card_rules": {
        "min_cards": 2,
        "target_cards": 5,
        "max_cards": 5,
        "strict_insufficiency_notice": True,
        "do_not_pad_placeholders": True,
    },
    "le_watchout_rules": {
        "enabled": True,
        "include_on_material_shift": True,
        "include_on_completeness_gap": True,
        "materiality": {
            "currency_mm": 1.0,
            "percent_pp": 0.5,
        },
    },
    "citation_rules": {
        "required_fields": ["path", "location", "excerpt"],
        "max_citations_per_card": 4,
        "excerpt_max_chars": 240,
    },
    "supplementary_rules": {
        "phase": "phase1",
        "include_current_pack_workbooks": True,
        "include_persistent_xlsx": True,
        "persistent_root": "data/context/persistent",
        "exclude_extensions": [".pdf"],
        "analysis_row_limit": 2500,
        "audit_min_score": 4,
    },
}

VARIANCE_QUESTION_PROMPT_VERSION = "variance_challenge_v1"
VARIANCE_QUESTION_PROMPT = (
    "Generate executive hot questions from tokenized workbook evidence.\n"
    "Rules:\n"
    "1. Every question must cite a named metric and at least one numeric delta.\n"
    "2. Use period bases explicitly: MoM, QoQ, YoY (or vs Budget).\n"
    "3. Prefer P&L and operating levers (Sales, AOI, EBITDA, SSS%, SST%).\n"
    "4. Ask for driver split, one-time vs structural mix, and next-month carryover.\n"
    "5. Include one concise prepared answer with the same metric deltas.\n"
    "6. Do not output generic questions without numbers."
)

VARIANCE_METRIC_SPECS: dict[str, dict[str, str]] = {
    "Total Sales": {"unit": "mm", "driver_focus": "traffic, ticket/mix, and calendar timing"},
    "AOI": {"unit": "mm", "driver_focus": "gross margin, labor, and controllable spend"},
    "Total EBITDA": {"unit": "mm", "driver_focus": "flow-through from sales plus fixed cost absorption"},
    "Franchise AOI": {"unit": "mm", "driver_focus": "royalty base, bad debt, and franchise fee timing"},
    "Property AOI": {"unit": "mm", "driver_focus": "rent, occupancy, and property-level cost mix"},
    "G&A AOI": {"unit": "mm", "driver_focus": "corporate cost timing and run-rate discipline"},
    "SSS%": {"unit": "pct", "driver_focus": "transactions versus check and promo mix"},
    "SST%": {"unit": "pct", "driver_focus": "traffic trend versus seasonality and competitive pressure"},
}

METRIC_KEYWORDS: dict[str, list[str]] = {
    "Total Sales": ["sales", "system sales", "total sales", "comp", "sss", "sst", "traffic", "ticket", "mix"],
    "AOI": ["aoi", "operating income", "margin", "flow-through", "pricing", "food", "paper", "labor"],
    "Total EBITDA": ["ebitda", "adjusted ebitda", "flow-through", "margin", "overhead", "g&a"],
    "Franchise AOI": ["franchise", "royalty", "fee", "bad debt", "successor", "franchise aoi"],
    "Property AOI": ["property", "rent", "occupancy", "utilities", "maintenance", "property aoi"],
    "G&A AOI": ["g&a", "corporate", "overhead", "incentive", "professional fees", "travel"],
    "SSS%": ["sss", "same store sales", "comps", "comp sales", "mix", "ticket"],
    "SST%": ["sst", "traffic", "transactions", "guest count", "frequency"],
}

SUPPLEMENTARY_SIGNAL_TOKENS = [
    "variance",
    "var",
    "budget",
    "le",
    "plan",
    "actual",
    "driver",
    "bridge",
    "commentary",
    "headwind",
    "tailwind",
]

SUPPLEMENTARY_SHEET_PRIORITIES = [
    "summary",
    "bridge",
    "check",
    "act x ple",
    "traffic",
    "sales",
    "aoi",
    "ebitda",
]

NARRATIVE_CUE_TOKENS = [
    "what worked",
    "didn't work",
    "drivers and impacts",
    "driven by",
    "due to",
    "because",
    "headwind",
    "tailwind",
    "traffic softness",
    "calendar initiatives",
    "weather",
    "timing",
    "consumer spending",
    "mix",
    "pricing",
    "labor",
    "commodity",
    "promotion",
    "promo",
]
NUMERIC_RE = re.compile(r"\(?-?\$?\d[\d,]*(?:\.\d+)?%?\)?")

BRIDGE_CUE_TOKENS = [
    "bridge",
    "variance to budget",
    "variance to le",
    "vs budget",
    "vs le",
    "vs py",
]

FOOTER_CUE_TOKENS = [
    "confidential and proprietary information of restaurant brands international",
    "source:",
]

REGION_TOKEN_MAP: dict[str, list[str]] = {
    "Canada": ["th can", "th canada", "canada", "ca "],
    "US": ["th us", " us ", "united states"],
    "C&US": ["c&us", "th c&us", "th c & us"],
}
VARIANCE_DRIVER_LEXICON: dict[str, list[str]] = {
    "traffic": ["traffic", "transactions", "guest count", "sst"],
    "check_mix": ["check", "mix", "ticket", "sss", "promo", "promotion"],
    "pricing": ["pricing", "price", "menu", "inflation"],
    "labor": ["labor", "wage", "staffing", "hours"],
    "commodity": ["commodity", "food", "paper", "cost of sales"],
    "media": ["media", "advertising", "digital", "campaign"],
    "weather": ["weather", "storm", "temperature"],
    "calendar": ["calendar", "timing", "one-time", "lap", "holiday"],
}


@dataclass
class Issue:
    location: str
    issue_type: str
    description: str
    severity: str
    recommended_fix: str
    evidence_refs: list[str]

    def as_dict(self) -> dict[str, Any]:
        return {
            "location": self.location,
            "issue_type": self.issue_type,
            "description": self.description,
            "severity": self.severity,
            "recommended_fix": self.recommended_fix,
            "evidence_refs": self.evidence_refs,
        }


def _iter_slide_files(pack_dir: Path) -> list[Path]:
    return sorted(pack_dir.glob("decks/*/slides/slide_*.json"))


def _iter_workbook_meta(pack_dir: Path) -> list[Path]:
    return sorted(pack_dir.glob("workbooks/*/workbook_meta.json"))


def _clamp_score(value: float) -> float:
    return max(0.0, min(100.0, value))


def _score_band(score_total: float, thresholds: dict[str, Any]) -> str:
    green_min = float(thresholds.get("green_min", 80))
    yellow_min = float(thresholds.get("yellow_min", 65))
    if score_total >= green_min:
        return "Green"
    if score_total >= yellow_min:
        return "Yellow"
    return "Red"


def _merge_dict(base: dict[str, Any], override: dict[str, Any] | None) -> dict[str, Any]:
    if not override:
        return dict(base)
    merged = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            nested = dict(merged[key])
            nested.update(value)
            merged[key] = nested
        else:
            merged[key] = value
    return merged


def load_hotq_scoring_config(config_path: Path | None = None) -> dict[str, Any]:
    if config_path and config_path.exists():
        user_config = read_json(config_path)
        config = _merge_dict(DEFAULT_HOTQ_SCORING_CONFIG, user_config)
    else:
        config = dict(DEFAULT_HOTQ_SCORING_CONFIG)
    weights = config.get("weights", {})
    total_weight = sum(float(weights.get(name, 0)) for name in DEFAULT_HOTQ_SCORING_CONFIG["weights"])
    if total_weight <= 0:
        raise ValueError("Invalid scoring config: sum of weights must be greater than zero.")
    return config


def load_hotq_policy(config_path: Path | None = None) -> dict[str, Any]:
    if config_path and config_path.exists():
        user_policy = read_json(config_path)
        policy = _merge_dict(DEFAULT_HOTQ_POLICY, user_policy)
    else:
        policy = dict(DEFAULT_HOTQ_POLICY)

    card_rules = policy.get("card_rules", {})
    min_cards = int(card_rules.get("min_cards", 2))
    target_cards = int(card_rules.get("target_cards", 5))
    max_cards = int(card_rules.get("max_cards", 5))
    if min_cards < 1:
        raise ValueError("Invalid policy: card_rules.min_cards must be >= 1.")
    if target_cards < min_cards:
        target_cards = min_cards
    if max_cards < target_cards:
        max_cards = target_cards
    card_rules["min_cards"] = min_cards
    card_rules["target_cards"] = target_cards
    card_rules["max_cards"] = max_cards
    policy["card_rules"] = card_rules

    policy_version = str(policy.get("policy_version", "")).strip()
    if not policy_version:
        raise ValueError("Invalid policy: policy_version is required.")

    return policy


def _infer_period_from_pack_dir(pack_dir: Path) -> str:
    parts = pack_dir.parts
    if len(parts) >= 2:
        return parts[-2]
    return "unknown-period"


def _infer_pack_type_from_pack_dir(pack_dir: Path) -> str:
    parts = pack_dir.parts
    if parts:
        return parts[-1]
    return "unknown-pack"


def _read_pack_summary(pack_dir: Path) -> dict[str, Any]:
    summary_path = pack_dir / "pack_summary.json"
    if summary_path.exists():
        return read_json(summary_path)
    return {}


def _parse_numeric_cell(value: str) -> float | None:
    cell = str(value).strip()
    if not cell or cell.startswith("="):
        return None
    try:
        return float(cell)
    except ValueError:
        return None


def _shift_period(period: str, delta: int) -> str | None:
    match = re.fullmatch(r"(\d{4})-P(\d{2})", period)
    if not match:
        return None
    year = int(match.group(1))
    month = int(match.group(2))
    month += delta
    while month <= 0:
        year -= 1
        month += 12
    while month > 12:
        year += 1
        month -= 12
    return f"{year}-P{month:02d}"


def _repo_root_from_pack_dir(pack_dir: Path) -> Path | None:
    for ancestor in [pack_dir, *pack_dir.parents]:
        if ancestor.name == "normalized" and ancestor.parent.name == "data":
            return ancestor.parent.parent
    return None


def _pick_variance_sheet_csv(pack_dir: Path) -> tuple[Path | None, str]:
    workbooks_dir = pack_dir / "workbooks"
    if not workbooks_dir.exists():
        return None, ""

    candidates: list[tuple[int, Path, str]] = []
    for workbook_dir in sorted([path for path in workbooks_dir.iterdir() if path.is_dir()]):
        name = workbook_dir.name.lower()
        priority = 0
        if "offline" in name:
            priority += 30
        if "close-template" in name or "preview-template" in name:
            priority += 20
        if "aoi-version" in name:
            priority += 10
        for sheet_slug in ["thca-p-l-aoi", "thca-p-l-summary"]:
            values_csv = workbook_dir / "sheets" / sheet_slug / "values.csv"
            if values_csv.exists():
                evidence_ref = str(values_csv.relative_to(pack_dir).as_posix())
                candidates.append((priority, values_csv, evidence_ref))

    if not candidates:
        return None, ""
    chosen = sorted(candidates, key=lambda item: item[0], reverse=True)[0]
    return chosen[1], chosen[2]


def _load_metric_snapshot(pack_dir: Path) -> tuple[dict[str, dict[str, Any]], str]:
    values_csv, evidence_ref = _pick_variance_sheet_csv(pack_dir)
    if values_csv is None:
        return {}, ""

    with values_csv.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))

    if len(rows) < 2:
        return {}, evidence_ref

    data_rows = rows[1:]
    base_idx: int | None = None
    header_row_idx: int | None = None
    for idx, row in enumerate(data_rows):
        lower = [str(cell).strip().lower() for cell in row]
        if "prior year" in lower and "actual" in lower and "budget" in lower:
            try:
                base_idx = lower.index("prior year")
                if base_idx + 3 < len(row):
                    header_row_idx = idx
                    break
            except ValueError:
                continue

    if base_idx is None or header_row_idx is None or base_idx <= 0:
        return {}, evidence_ref

    snapshot: dict[str, dict[str, Any]] = {}
    for row in data_rows[header_row_idx + 1 :]:
        if len(row) <= base_idx + 3:
            continue
        metric_label = str(row[base_idx - 1]).strip()
        if metric_label not in VARIANCE_METRIC_SPECS:
            continue
        snapshot[metric_label] = {
            "row_number": str(row[0]).strip(),
            "prior_year": _parse_numeric_cell(row[base_idx]),
            "actual": _parse_numeric_cell(row[base_idx + 1]),
            "le": _parse_numeric_cell(row[base_idx + 2]),
            "budget": _parse_numeric_cell(row[base_idx + 3]),
            "evidence_ref": evidence_ref,
        }
    return snapshot, evidence_ref


def _format_signed_mm(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"{value:+.1f}MM"


def _format_pct(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"{value * 100:.2f}%"


def _format_signed_pp(delta: float | None) -> str:
    if delta is None:
        return "n/a"
    return f"{delta * 100:+.2f}pp"


def _format_signed_pct(delta: float | None) -> str:
    if delta is None:
        return "n/a"
    return f"{delta:+.1f}%"


def _safe_percent_delta(current: float | None, baseline: float | None) -> float | None:
    if current is None or baseline is None or abs(baseline) < 1e-9:
        return None
    return (current - baseline) / abs(baseline) * 100.0


def _load_manifest_roles(pack_dir: Path) -> dict[str, str]:
    manifest_path = pack_dir / "pack_manifest.json"
    if not manifest_path.exists():
        return {}
    manifest = read_json(manifest_path)
    role_map: dict[str, str] = {}
    for item in manifest.get("files", []):
        slug = str(item.get("file_slug", "")).strip()
        role = str(item.get("role", "")).strip()
        if slug and role:
            role_map[slug] = role
    return role_map


def _sheet_priority(sheet_name: str) -> int:
    lowered = sheet_name.lower()
    score = 0
    for index, token in enumerate(SUPPLEMENTARY_SHEET_PRIORITIES):
        if token in lowered:
            score += 20 - index
    return score


def _format_numeric_for_snippet(value: float) -> str:
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    if abs(value) >= 10:
        return f"{value:.1f}"
    return f"{value:.3f}"


def _collect_persistent_xlsx_evidence(
    *,
    pack_dir: Path,
    metric_targets: list[str],
    policy: dict[str, Any],
) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    snippets_by_metric: dict[str, list[dict[str, Any]]] = {metric: [] for metric in metric_targets}
    audit_notes: list[dict[str, Any]] = []
    supplementary_rules = policy.get("supplementary_rules", {})
    if not bool(supplementary_rules.get("include_persistent_xlsx", True)):
        return snippets_by_metric, audit_notes

    repo_root = _repo_root_from_pack_dir(pack_dir)
    if not repo_root:
        audit_notes.append(
            {
                "scope": "supplementary_persistent",
                "status": "not_helpful",
                "reason": "Unable to resolve repo root from pack directory.",
            }
        )
        return snippets_by_metric, audit_notes

    persistent_root = repo_root / str(supplementary_rules.get("persistent_root", "data/context/persistent"))
    if not persistent_root.exists():
        audit_notes.append(
            {
                "scope": "supplementary_persistent",
                "status": "not_helpful",
                "reason": f"Persistent root not found: {persistent_root.relative_to(repo_root).as_posix()}",
            }
        )
        return snippets_by_metric, audit_notes

    try:
        import openpyxl  # type: ignore
    except Exception:
        audit_notes.append(
            {
                "scope": "supplementary_persistent",
                "status": "not_helpful",
                "reason": "openpyxl dependency unavailable; persistent XLSX scan skipped.",
            }
        )
        return snippets_by_metric, audit_notes

    excluded_ext = {str(ext).lower() for ext in supplementary_rules.get("exclude_extensions", [".pdf"])}
    row_limit = int(supplementary_rules.get("analysis_row_limit", 2500))
    xlsx_files = sorted(path for path in persistent_root.rglob("*.xlsx") if path.suffix.lower() not in excluded_ext)
    if not xlsx_files:
        audit_notes.append(
            {
                "scope": "supplementary_persistent",
                "status": "not_helpful",
                "reason": "No persistent XLSX files found in scope.",
            }
        )
        return snippets_by_metric, audit_notes

    files_scanned = 0
    for workbook_path in xlsx_files:
        files_scanned += 1
        try:
            workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        except Exception:
            continue
        workbook_slug = workbook_path.stem
        rel_path = str(workbook_path.relative_to(repo_root).as_posix())
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            metric_hits = 0
            max_row = int(sheet.max_row or row_limit)
            for row_idx, row_values in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(row_limit, max_row), values_only=True),
                start=1,
            ):
                row_text_parts = [str(cell).strip() for cell in row_values if cell is not None and str(cell).strip()]
                if not row_text_parts:
                    continue
                joined = " ".join(row_text_parts).lower()
                numeric_mentions = NUMERIC_RE.findall(joined)
                if not numeric_mentions:
                    continue
                signal_hits = sum(1 for token in SUPPLEMENTARY_SIGNAL_TOKENS if token in joined)
                if signal_hits == 0:
                    continue
                for metric in metric_targets:
                    if len(snippets_by_metric[metric]) >= 4:
                        continue
                    keyword_hits = [token for token in METRIC_KEYWORDS.get(metric, []) if token in joined]
                    if not keyword_hits:
                        continue
                    score = len(keyword_hits) * 3 + signal_hits + min(2, len(numeric_mentions))
                    snippets_by_metric[metric].append(
                        {
                            "metric": metric,
                            "workbook_slug": workbook_slug,
                            "sheet_name": sheet_name,
                            "row_number": str(row_idx),
                            "matched_keywords": keyword_hits[:4],
                            "numeric_values": numeric_mentions[:4],
                            "snippet_text": " | ".join(row_text_parts[:6])[:220],
                            "score": score,
                            "evidence_ref": rel_path,
                        }
                    )
                    metric_hits += 1
                if metric_hits >= 8:
                    break
        workbook.close()

    audit_notes.append(
        {
            "scope": "supplementary_persistent",
            "status": "evaluated",
            "reason": f"Persistent XLSX scan completed; files_scanned={files_scanned}.",
        }
    )
    return snippets_by_metric, audit_notes


def _collect_supplementary_evidence(
    pack_dir: Path,
    ranked_candidates: list[dict[str, Any]],
    policy: dict[str, Any],
) -> tuple[dict[str, list[dict[str, Any]]], list[str], list[dict[str, Any]]]:
    supplementary_rules = policy.get("supplementary_rules", {})
    metric_targets = [str(item.get("metric", "")) for item in ranked_candidates[:6] if item.get("metric")]
    if not metric_targets:
        return {}, [], []

    snippets_by_metric: dict[str, list[dict[str, Any]]] = {metric: [] for metric in metric_targets}
    evidence_refs: list[str] = []
    evidence_gap_registry: list[dict[str, Any]] = []
    include_current_pack = bool(supplementary_rules.get("include_current_pack_workbooks", True))
    workbooks_dir = pack_dir / "workbooks"

    if include_current_pack and workbooks_dir.exists():
        role_map = _load_manifest_roles(pack_dir)
        primary_sheet_csv, _ = _pick_variance_sheet_csv(pack_dir)
        primary_workbook = primary_sheet_csv.parents[2].name if primary_sheet_csv else ""
        workbook_dirs = sorted(path for path in workbooks_dir.iterdir() if path.is_dir())
        for workbook_dir in workbook_dirs:
            workbook_slug = workbook_dir.name
            role = role_map.get(workbook_slug, "")
            if workbook_slug == primary_workbook:
                continue
            if role and role != "supporting_excel":
                continue

            meta_path = workbook_dir / "workbook_meta.json"
            if not meta_path.exists():
                continue
            meta = read_json(meta_path)
            sheets = sorted(
                meta.get("sheets", []),
                key=lambda item: _sheet_priority(str(item.get("sheet_name", ""))),
                reverse=True,
            )
            for sheet in sheets:
                values_csv_rel = str(sheet.get("values_csv", ""))
                if not values_csv_rel:
                    continue
                values_csv = workbook_dir / values_csv_rel
                if not values_csv.exists():
                    continue

                sheet_name = str(sheet.get("sheet_name", ""))
                row_limit = int(supplementary_rules.get("analysis_row_limit", 2500))
                if _sheet_priority(sheet_name) == 0:
                    row_limit = min(row_limit, 600)
                metric_hits = 0
                with values_csv.open("r", encoding="utf-8", newline="") as handle:
                    reader = csv.reader(handle)
                    next(reader, None)
                    for row_idx, row in enumerate(reader, start=1):
                        if row_idx > row_limit:
                            break
                        row_text_parts = [str(cell).strip() for cell in row[1:] if str(cell).strip()]
                        if not row_text_parts:
                            continue
                        joined = " ".join(row_text_parts).lower()
                        numeric_values = [value for value in (_parse_numeric_cell(cell) for cell in row[1:]) if value is not None]
                        if not numeric_values:
                            continue
                        signal_hits = sum(1 for token in SUPPLEMENTARY_SIGNAL_TOKENS if token in joined)
                        if signal_hits == 0:
                            continue

                        for metric in metric_targets:
                            if len(snippets_by_metric[metric]) >= 4:
                                continue
                            keyword_hits = [token for token in METRIC_KEYWORDS.get(metric, []) if token in joined]
                            if not keyword_hits:
                                continue
                            score = len(keyword_hits) * 3 + signal_hits
                            snippet = {
                                "metric": metric,
                                "workbook_slug": workbook_slug,
                                "sheet_name": sheet_name,
                                "row_number": str(row[0]).strip() if row else str(row_idx),
                                "matched_keywords": keyword_hits[:4],
                                "numeric_values": [_format_numeric_for_snippet(value) for value in numeric_values[:4]],
                                "snippet_text": " | ".join(row_text_parts[:6])[:220],
                                "score": score,
                                "evidence_ref": str(values_csv.relative_to(pack_dir).as_posix()),
                            }
                            snippets_by_metric[metric].append(snippet)
                            evidence_refs.append(snippet["evidence_ref"])
                            metric_hits += 1
                        if metric_hits >= 8:
                            break
    else:
        evidence_gap_registry.append(
            {
                "scope": "supplementary_current_pack",
                "status": "not_helpful",
                "reason": "Current pack workbook scope unavailable for supplementary scan.",
            }
        )

    persistent_snippets, persistent_notes = _collect_persistent_xlsx_evidence(
        pack_dir=pack_dir,
        metric_targets=metric_targets,
        policy=policy,
    )
    evidence_gap_registry.extend(persistent_notes)
    for metric, snippets in persistent_snippets.items():
        for snippet in snippets:
            snippets_by_metric[metric].append(snippet)
            evidence_refs.append(str(snippet.get("evidence_ref", "")))

    audit_min_score = int(supplementary_rules.get("audit_min_score", 4))
    for metric in metric_targets:
        snippets_by_metric[metric] = sorted(
            snippets_by_metric[metric],
            key=lambda item: float(item.get("score", 0)),
            reverse=True,
        )[:3]
        if not snippets_by_metric[metric]:
            evidence_gap_registry.append(
                {
                    "scope": "supplementary_metric_audit",
                    "metric": metric,
                    "status": "not_helpful",
                    "reason": "No supplementary workbook evidence met signal thresholds.",
                }
            )
            continue
        top_score = float(snippets_by_metric[metric][0].get("score", 0.0))
        if top_score < audit_min_score:
            evidence_gap_registry.append(
                {
                    "scope": "supplementary_metric_audit",
                    "metric": metric,
                    "status": "not_helpful",
                    "reason": f"Supplementary evidence scored below audit threshold ({top_score:.1f}<{audit_min_score}).",
                }
            )
        else:
            evidence_gap_registry.append(
                {
                    "scope": "supplementary_metric_audit",
                    "metric": metric,
                    "status": "helpful",
                    "reason": f"Supplementary evidence passed audit threshold ({top_score:.1f}).",
                }
            )

    deduped_evidence_refs = sorted(ref for ref in set(evidence_refs) if ref)
    return snippets_by_metric, deduped_evidence_refs, evidence_gap_registry


def _compute_le_change_flags(
    current_snapshot: dict[str, dict[str, Any]],
    previous_snapshot: dict[str, dict[str, Any]],
    policy: dict[str, Any],
) -> list[dict[str, Any]]:
    flags: list[dict[str, Any]] = []
    le_materiality = policy.get("le_watchout_rules", {}).get("materiality", {})
    mm_threshold = float(le_materiality.get("currency_mm", 1.0))
    pct_threshold = float(le_materiality.get("percent_pp", 0.5)) / 100.0
    for metric, spec in VARIANCE_METRIC_SPECS.items():
        current = current_snapshot.get(metric, {})
        previous = previous_snapshot.get(metric, {})
        current_le_raw = current.get("le")
        previous_le_raw = previous.get("le")
        current_le = (
            float(current_le_raw)
            if isinstance(current_le_raw, (int, float)) and abs(float(current_le_raw)) > 1e-9
            else None
        )
        previous_le = (
            float(previous_le_raw)
            if isinstance(previous_le_raw, (int, float)) and abs(float(previous_le_raw)) > 1e-9
            else None
        )
        if current_le is None and previous_le is not None:
            flags.append(
                {
                    "metric": metric,
                    "current_le": current_le_raw,
                    "previous_le": previous_le_raw,
                    "le_change": None,
                    "le_change_pct": None,
                    "status": "missing_current_period_le",
                }
            )
            continue
        if current_le is None or previous_le is None:
            continue
        delta = float(current_le) - float(previous_le)
        if spec["unit"] == "pct":
            material = abs(delta) >= pct_threshold
        else:
            material = abs(delta) >= mm_threshold
        if not material:
            continue
        flags.append(
            {
                "metric": metric,
                "current_le": current_le,
                "previous_le": previous_le,
                "le_change": delta,
                "le_change_pct": _safe_percent_delta(current_le, previous_le)
                if spec["unit"] == "mm"
                else None,
                "status": "favorable" if delta >= 0 else "unfavorable",
            }
        )
    return flags


def _format_basis_delta(item: dict[str, Any], *, basis: str, unit: str) -> str:
    value = item.get(basis)
    if unit == "pct":
        return _format_signed_pp(value)
    return _format_signed_mm(value)


def _basis_phrase(item: dict[str, Any], *, basis: str, unit: str) -> str:
    label = "vs Budget" if basis == "vs_budget" else "vs LE"
    value = item.get(basis)
    if value is None:
        if basis == "vs_le":
            return "LE not populated"
        return f"n/a {label}"
    return f"{_format_basis_delta(item, basis=basis, unit=unit)} {label}"


def _region_from_text(text: str) -> str:
    lowered = text.lower()
    if any(token in lowered for token in REGION_TOKEN_MAP["C&US"]):
        return "C&US"
    has_canada = any(token in lowered for token in REGION_TOKEN_MAP["Canada"])
    has_us = any(token in lowered for token in REGION_TOKEN_MAP["US"])
    if has_canada and has_us:
        return "C&US"
    if has_canada:
        return "Canada"
    if has_us:
        return "US"
    return "C&US"


def _infer_narrative_block_class(text: str, numeric_density: float, line_count: int) -> str:
    lowered = text.lower()
    if any(token in lowered for token in FOOTER_CUE_TOKENS):
        return "footer"
    if any(token in lowered for token in BRIDGE_CUE_TOKENS):
        return "bridge_summary"
    if numeric_density >= 1.3 or (line_count >= 10 and numeric_density > 0.6):
        return "table_like"
    if any(token in lowered for token in NARRATIVE_CUE_TOKENS):
        return "narrative"
    if len(lowered) >= 120 and numeric_density < 0.8:
        return "narrative"
    return "table_like"


def _narrative_score(text: str, numeric_density: float, block_class: str) -> float:
    lowered = text.lower()
    cue_hits = sum(1 for token in NARRATIVE_CUE_TOKENS if token in lowered)
    bridge_hits = sum(1 for token in BRIDGE_CUE_TOKENS if token in lowered)
    base = cue_hits * 2.2 - bridge_hits * 1.5 - numeric_density * 1.2
    if block_class == "narrative":
        base += 3.0
    if block_class == "footer":
        base -= 5.0
    return round(base, 3)


def _collect_narrative_blocks(pack_dir: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    class_counts = {"narrative": 0, "bridge_summary": 0, "table_like": 0, "footer": 0}
    region_counts = {"C&US": 0, "Canada": 0, "US": 0}
    for slide_path in _iter_slide_files(pack_dir):
        slide = read_json(slide_path)
        slide_number = int(slide.get("slide_number", 0) or 0)
        slide_title = str(slide.get("title", "")).strip()
        evidence_ref = str(slide_path.relative_to(pack_dir).as_posix())
        block_list = slide.get("text_blocks")
        normalized_blocks: list[dict[str, Any]] = []
        if isinstance(block_list, list) and block_list:
            for block in block_list:
                lines = [str(item).strip() for item in block.get("lines", []) if str(item).strip()]
                if not lines:
                    continue
                joined = " ".join(lines)
                numeric_density = float(block.get("numeric_density", 0.0))
                block_class = str(block.get("block_class", "")).strip() or _infer_narrative_block_class(
                    joined, numeric_density=numeric_density, line_count=len(lines)
                )
                signal_score = float(block.get("narrative_signal_score", _narrative_score(joined, numeric_density, block_class)))
                normalized_blocks.append(
                    {
                        "block_index": int(block.get("block_index", len(normalized_blocks) + 1)),
                        "lines": lines,
                        "text": joined,
                        "char_count": int(block.get("char_count", len(joined))),
                        "numeric_density": numeric_density,
                        "block_class": block_class,
                        "narrative_signal_score": signal_score,
                    }
                )
        if not normalized_blocks:
            fallback_lines = [slide_title, *[str(item) for item in slide.get("body", [])], str(slide.get("note_text", ""))]
            fallback_lines = [line.strip() for line in fallback_lines if line and str(line).strip()]
            if fallback_lines:
                joined = " ".join(fallback_lines)
                numeric_mentions = len(NUMERIC_RE.findall(joined))
                numeric_density = numeric_mentions / max(1, len(fallback_lines))
                block_class = _infer_narrative_block_class(joined, numeric_density=numeric_density, line_count=len(fallback_lines))
                normalized_blocks.append(
                    {
                        "block_index": 1,
                        "lines": fallback_lines,
                        "text": joined,
                        "char_count": len(joined),
                        "numeric_density": numeric_density,
                        "block_class": block_class,
                        "narrative_signal_score": _narrative_score(joined, numeric_density, block_class),
                    }
                )

        for block in normalized_blocks:
            text = str(block["text"])
            region = _region_from_text(f"{slide_title} {text}")
            class_counts[block["block_class"]] = class_counts.get(block["block_class"], 0) + 1
            region_counts[region] = region_counts.get(region, 0) + 1
            blocks.append(
                {
                    "slide_number": slide_number,
                    "slide_title": slide_title,
                    "evidence_ref": evidence_ref,
                    "region": region,
                    **block,
                }
            )

    summary = {
        "total_blocks": len(blocks),
        "class_counts": class_counts,
        "region_counts": region_counts,
        "narrative_blocks": sum(1 for block in blocks if block.get("block_class") == "narrative"),
    }
    return blocks, summary


def _classify_scope(text: str, policy: dict[str, Any]) -> tuple[str, bool, int, int]:
    scope_rules = policy.get("scope_rules", {})
    lowered = text.lower()
    restaurant_tokens = [str(token).lower() for token in scope_rules.get("restaurant_tokens", [])]
    cpg_tokens = [str(token).lower() for token in scope_rules.get("cpg_tokens", [])]
    restaurant_hits = sum(1 for token in restaurant_tokens if token and token in lowered)
    cpg_hits = sum(1 for token in cpg_tokens if token and token in lowered)
    quantified_regex = str(scope_rules.get("quantified_bridge_regex", ""))
    bridge_quantified = bool(quantified_regex and re.search(quantified_regex, lowered))
    if cpg_hits > restaurant_hits and not bridge_quantified:
        return "cpg_primary", bridge_quantified, restaurant_hits, cpg_hits
    if cpg_hits and restaurant_hits and bridge_quantified:
        return "quantified_bridge", bridge_quantified, restaurant_hits, cpg_hits
    if restaurant_hits and cpg_hits:
        return "restaurant_mixed", bridge_quantified, restaurant_hits, cpg_hits
    if restaurant_hits:
        return "restaurant_primary", bridge_quantified, restaurant_hits, cpg_hits
    if cpg_hits:
        return "cpg_secondary", bridge_quantified, restaurant_hits, cpg_hits
    return "unclassified", bridge_quantified, restaurant_hits, cpg_hits


def _metric_narrative_matches(metric: str, blocks: list[dict[str, Any]], policy: dict[str, Any]) -> list[dict[str, Any]]:
    keywords = METRIC_KEYWORDS.get(metric, [])
    driver_tokens = [token for tokens in VARIANCE_DRIVER_LEXICON.values() for token in tokens]
    matches: list[dict[str, Any]] = []
    for block in blocks:
        text = str(block.get("text", "")).lower()
        keyword_hits = [token for token in keywords if token in text]
        driver_hits = [token for token in driver_tokens if token in text]
        if not keyword_hits and not driver_hits:
            continue
        slide_title = str(block.get("slide_title", "")).lower()
        score = float(block.get("narrative_signal_score", 0.0)) + len(keyword_hits) * 2.0 + len(driver_hits) * 1.5
        if any(token in slide_title for token in ("what worked", "didn't work", "drivers and impacts")):
            score += 6.0
        if "bridge" in slide_title:
            score -= 4.0
        block_class = str(block.get("block_class", ""))
        if block_class == "narrative":
            score += 3.0
        elif block_class == "bridge_summary":
            score -= 3.0
        elif block_class == "table_like":
            score -= 1.0
        scope_classification, bridge_quantified, restaurant_hit_count, cpg_hit_count = _classify_scope(text, policy)
        matches.append(
            {
                "evidence_ref": str(block.get("evidence_ref", "")),
                "slide_number": int(block.get("slide_number", 0)),
                "slide_title": str(block.get("slide_title", "")),
                "region": str(block.get("region", "C&US")),
                "block_class": block_class,
                "snippet": " | ".join(block.get("lines", [])[:3])[:280],
                "matched_driver_tokens": sorted(set(driver_hits))[:6],
                "scope_classification": scope_classification,
                "bridge_quantified": bridge_quantified,
                "restaurant_hit_count": restaurant_hit_count,
                "cpg_hit_count": cpg_hit_count,
                "score": round(score, 3),
            }
        )
    return sorted(matches, key=lambda item: float(item.get("score", 0.0)), reverse=True)


def _format_basis_summary(item: dict[str, Any], unit: str) -> dict[str, str]:
    if unit == "pct":
        return {
            "vs_budget": _format_signed_pp(item.get("vs_budget")),
            "vs_le": _format_signed_pp(item.get("vs_le")) if item.get("vs_le") is not None else "LE not populated",
            "mom": _format_signed_pp(item.get("mom_delta")),
            "qoq": _format_signed_pp(item.get("qoq_delta")),
            "yoy": _format_signed_pp(item.get("vs_py")),
        }
    return {
        "vs_budget": _format_signed_mm(item.get("vs_budget")),
        "vs_le": _format_signed_mm(item.get("vs_le")) if item.get("vs_le") is not None else "LE not populated",
        "mom": _format_signed_mm(item.get("mom_delta")),
        "qoq": _format_signed_mm(item.get("qoq_delta")),
        "yoy": _format_signed_mm(item.get("vs_py")),
    }


def _ensure_basis_presence(summary: dict[str, str]) -> bool:
    budget = summary.get("vs_budget", "").lower()
    le = summary.get("vs_le", "").lower()
    has_budget = budget not in {"", "n/a", "n/a vs budget"}
    has_le = le not in {"", "n/a", "le not populated"}
    return has_budget or has_le


def _build_quality_gate(challenge_cards: list[dict[str, Any]]) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []
    if not challenge_cards:
        return {
            "status": "fail",
            "checks": [
                {
                    "check": "challenge_cards_present",
                    "status": "fail",
                    "message": "No challenge cards were generated from the available pack evidence.",
                }
            ],
        }

    missing_narrative = [
        card.get("metric", "")
        for card in challenge_cards
        if card.get("card_type") != "le_watchout" and not card.get("narrative_evidence_refs")
    ]
    checks.append(
        {
            "check": "narrative_evidence_per_card",
            "status": "pass" if not missing_narrative else "fail",
            "message": "All non-watchout cards include narrative evidence."
            if not missing_narrative
            else f"Missing narrative evidence for: {', '.join(missing_narrative)}",
        }
    )

    missing_basis = [
        card.get("metric", "")
        for card in challenge_cards
        if card.get("card_type") != "le_watchout" and not _ensure_basis_presence(card.get("basis_summary", {}))
    ]
    checks.append(
        {
            "check": "basis_delta_presence",
            "status": "pass" if not missing_basis else "fail",
            "message": "All non-watchout cards include Budget/LE basis deltas."
            if not missing_basis
            else f"Missing basis deltas for: {', '.join(missing_basis)}",
        }
    )

    bridge_only = [
        card.get("metric", "")
        for card in challenge_cards
        if card.get("card_type") != "le_watchout"
        and card.get("narrative_block_classes")
        and all(cls in {"bridge_summary", "table_like"} for cls in card.get("narrative_block_classes", []))
    ]
    checks.append(
        {
            "check": "no_bridge_only_causal_claim",
            "status": "pass" if not bridge_only else "fail",
            "message": "No cards rely only on bridge/table blocks for causal narrative."
            if not bridge_only
            else f"Bridge-only causal support detected for: {', '.join(bridge_only)}",
        }
    )

    citation_missing = []
    for card in challenge_cards:
        if card.get("card_type") != "le_watchout":
            citations = card.get("citation_bundle", [])
            valid = False
            if isinstance(citations, list):
                for citation in citations:
                    if not isinstance(citation, dict):
                        continue
                    path = str(citation.get("path", "")).strip()
                    location = str(citation.get("location", "")).strip()
                    excerpt = str(citation.get("excerpt", "")).strip()
                    if path and location and excerpt and len(excerpt) >= 12:
                        valid = True
                        break
            if not valid:
                citation_missing.append(str(card.get("metric", "")))
    checks.append(
        {
            "check": "citation_bundle_per_card",
            "status": "pass" if not citation_missing else "fail",
            "message": "All non-watchout cards include citation bundle entries."
            if not citation_missing
            else f"Missing citation bundle for: {', '.join(citation_missing)}",
        }
    )

    failed = [check for check in checks if check["status"] == "fail"]
    status = "pass"
    if failed:
        status = "downgraded_narrative_gap"
    if len(failed) >= 2:
        status = "fail"
    return {"status": status, "checks": checks}


def _apply_scope_filters(
    matches: list[dict[str, Any]],
    *,
    metric: str,
    policy: dict[str, Any],
    scope_filters_applied: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    scope_rules = policy.get("scope_rules", {})
    filter_cpg_primary = bool(scope_rules.get("exclude_cpg_primary_without_bridge", True))
    filtered: list[dict[str, Any]] = []
    for match in matches:
        scope_classification = str(match.get("scope_classification", "unclassified"))
        bridge_quantified = bool(match.get("bridge_quantified", False))
        if filter_cpg_primary and scope_classification == "cpg_primary" and not bridge_quantified:
            scope_filters_applied.append(
                {
                    "metric": metric,
                    "evidence_ref": str(match.get("evidence_ref", "")),
                    "scope_classification": scope_classification,
                    "reason": "Excluded CPG-primary narrative without quantified restaurant bridge.",
                }
            )
            continue
        filtered.append(match)
    return filtered


def _citation_excerpt(text: str, max_chars: int) -> str:
    compact = " ".join(str(text).split())
    if len(compact) <= max_chars:
        return compact
    return compact[: max_chars - 3].rstrip() + "..."


def _build_citation_bundle(
    *,
    item: dict[str, Any],
    narrative_matches: list[dict[str, Any]],
    supplementary_snippets: list[dict[str, Any]],
    policy: dict[str, Any],
) -> list[dict[str, str]]:
    citation_rules = policy.get("citation_rules", {})
    max_citations = int(citation_rules.get("max_citations_per_card", 4))
    excerpt_max_chars = int(citation_rules.get("excerpt_max_chars", 240))
    bundle: list[dict[str, str]] = []

    for match in narrative_matches:
        path = str(match.get("evidence_ref", "")).strip()
        if not path:
            continue
        slide_number = int(match.get("slide_number", 0))
        location = f"slide {slide_number}" if slide_number > 0 else "slide"
        excerpt = _citation_excerpt(str(match.get("snippet", "")), excerpt_max_chars)
        if excerpt:
            bundle.append({"path": path, "location": location, "excerpt": excerpt})

    for snippet in supplementary_snippets:
        path = str(snippet.get("evidence_ref", "")).strip()
        if not path:
            continue
        sheet_name = str(snippet.get("sheet_name", "")).strip() or "sheet"
        row_number = str(snippet.get("row_number", "")).strip() or "n/a"
        location = f"{sheet_name} row {row_number}"
        excerpt = _citation_excerpt(str(snippet.get("snippet_text", "")), excerpt_max_chars)
        if excerpt:
            bundle.append({"path": path, "location": location, "excerpt": excerpt})

    base_path = str(item.get("evidence_ref", "")).strip()
    if base_path:
        metric = str(item.get("metric", "")).strip() or "metric"
        location = f"row {item.get('row_number', 'n/a')}"
        unit = str(item.get("unit", "mm"))
        basis_vs_budget = _format_basis_delta(item, basis="vs_budget", unit=unit)
        basis_vs_le = _format_basis_delta(item, basis="vs_le", unit=unit) if item.get("vs_le") is not None else "LE not populated"
        excerpt = _citation_excerpt(f"{metric}: {basis_vs_budget} vs Budget, {basis_vs_le} vs LE.", excerpt_max_chars)
        bundle.append({"path": base_path, "location": location, "excerpt": excerpt})

    deduped: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for citation in bundle:
        key = (citation["path"], citation["location"], citation["excerpt"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(citation)
        if len(deduped) >= max_citations:
            break
    return deduped


def _apply_term_guard_to_text(
    text: str,
    *,
    card_metric: str,
    field: str,
    policy: dict[str, Any],
) -> tuple[str, list[dict[str, Any]], bool]:
    term_guard = policy.get("term_guard", {})
    allowlist = [str(item).lower() for item in term_guard.get("allowlist", [])]
    banned_terms = term_guard.get("banned_terms", [])
    rewritten = text
    hits: list[dict[str, Any]] = []
    downgraded = False

    for rule in banned_terms:
        pattern = str(rule.get("pattern", "")).strip()
        if not pattern:
            continue
        regex = re.compile(pattern, re.IGNORECASE)
        matched = regex.search(rewritten)
        if not matched:
            continue
        matched_text = matched.group(0)
        lowered = rewritten.lower()
        if any(phrase in lowered and matched_text.lower() in phrase for phrase in allowlist):
            continue
        if bool(rule.get("allow_if_numeric")) and NUMERIC_RE.search(rewritten):
            continue

        replacement = str(rule.get("rewrite", "")).strip()
        if replacement:
            rewritten, replacements = regex.subn(replacement, rewritten)
            if replacements > 0:
                hits.append(
                    {
                        "metric": card_metric,
                        "field": field,
                        "term_id": str(rule.get("id", "unknown")),
                        "matched_text": matched_text,
                        "rewrite_applied": replacement,
                        "status": "rewritten",
                        "count": replacements,
                    }
                )
            continue

        rewritten, replacements = regex.subn("[unsupported wording removed]", rewritten)
        if replacements > 0:
            hits.append(
                {
                    "metric": card_metric,
                    "field": field,
                    "term_id": str(rule.get("id", "unknown")),
                    "matched_text": matched_text,
                    "rewrite_applied": "",
                    "status": "blocked",
                    "count": replacements,
                }
            )
            downgraded = True

    return rewritten, hits, downgraded


def _apply_term_guard_to_cards(
    cards: list[dict[str, Any]],
    *,
    policy: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    guarded: list[dict[str, Any]] = []
    term_guard_hits: list[dict[str, Any]] = []
    for card in cards:
        card_copy = dict(card)
        metric = str(card_copy.get("metric", "metric"))
        question, question_hits, question_downgraded = _apply_term_guard_to_text(
            str(card_copy.get("challenge_question", "")),
            card_metric=metric,
            field="challenge_question",
            policy=policy,
        )
        answer, answer_hits, answer_downgraded = _apply_term_guard_to_text(
            str(card_copy.get("prepared_answer", "")),
            card_metric=metric,
            field="prepared_answer",
            policy=policy,
        )
        card_copy["challenge_question"] = question
        card_copy["prepared_answer"] = answer
        if question_downgraded or answer_downgraded:
            card_copy["confidence"] = "low"
            card_copy["verify_next"] = (
                f"{card_copy.get('verify_next', '').strip()} "
                "Replace blocked shorthand with deck-native drivers before executive review."
            ).strip()
        term_guard_hits.extend(question_hits)
        term_guard_hits.extend(answer_hits)
        guarded.append(card_copy)
    return guarded, term_guard_hits


def _primary_basis(item: dict[str, Any], unit: str) -> tuple[str, str]:
    priority = "vs_budget"
    if item.get("vs_le") is not None and abs(float(item.get("vs_le") or 0.0)) >= abs(float(item.get("vs_budget") or 0.0)):
        priority = "vs_le"
    secondary = "vs_le" if priority == "vs_budget" else "vs_budget"
    return _basis_phrase(item, basis=priority, unit=unit), _basis_phrase(item, basis=secondary, unit=unit)


def _build_challenge_card(
    *,
    item: dict[str, Any],
    period: str,
    region: str,
    narrative_matches: list[dict[str, Any]],
    supplementary_snippets: list[dict[str, Any]],
    policy: dict[str, Any],
) -> dict[str, Any]:
    metric = str(item["metric"])
    unit = str(item["unit"])
    primary_basis, secondary_basis = _primary_basis(item, unit)
    basis_summary = _format_basis_summary(item, unit)
    narrative_refs = [str(match.get("evidence_ref", "")) for match in narrative_matches if match.get("evidence_ref")]
    narrative_classes = [str(match.get("block_class", "")) for match in narrative_matches if match.get("block_class")]
    supplementary_refs = [str(snippet.get("evidence_ref", "")) for snippet in supplementary_snippets if snippet.get("evidence_ref")]

    mom = _format_signed_pp(item.get("mom_delta")) if unit == "pct" else _format_signed_mm(item.get("mom_delta"))
    qoq = _format_signed_pp(item.get("qoq_delta")) if unit == "pct" else _format_signed_mm(item.get("qoq_delta"))
    yoy = _format_signed_pp(item.get("vs_py")) if unit == "pct" else _format_signed_mm(item.get("vs_py"))
    narrative_text = (
        str(narrative_matches[0].get("snippet", ""))
        if narrative_matches
        else "Narrative commentary support is thin in extracted slides."
    )
    supplementary_text = ""
    if supplementary_snippets:
        snippet = supplementary_snippets[0]
        supplementary_text = (
            f" Supporting workbook evidence ({snippet.get('workbook_slug')}:{snippet.get('sheet_name')} "
            f"row {snippet.get('row_number')}): {snippet.get('snippet_text')}."
        )

    challenge_question = (
        f"{region} {metric} challenge: {primary_basis} in {period} with {mom} MoM, {qoq} QoQ, and {yoy} YoY. "
        "Which driver split best explains the variance versus Budget/LE and what should reverse next month?"
    )
    prepared_answer = (
        f"{metric} is anchored at {primary_basis} (secondary: {secondary_basis}). "
        f"Deck narrative points to: {narrative_text}. Focus defense on {item.get('driver_focus', '')}."
        f"{supplementary_text}"
    )
    why_now = (
        f"Budget/LE pressure is material in {period} and requires a causal storyline before executive review."
    )
    if item.get("le_change_vs_prior_month") is not None:
        le_delta = float(item.get("le_change_vs_prior_month") or 0.0)
        le_label = _format_signed_pp(le_delta) if unit == "pct" else _format_signed_mm(le_delta)
        le_change_label = str(policy.get("semantics", {}).get("le_change_label", "vs prior LE"))
        why_now = f"{why_now} LE moved {le_label} {le_change_label}."

    verify_next = (
        "Confirm bridge tie-out to row "
        f"{item.get('row_number', 'n/a')} and validate one-time vs structural split in commentary notes."
    )
    if supplementary_snippets:
        verify_next = (
            f"{verify_next} Cross-check supporting workbook {supplementary_snippets[0].get('workbook_slug')} for driver proof."
        )

    confidence = "high"
    if not narrative_refs:
        confidence = "low"
    elif not _ensure_basis_presence(basis_summary):
        confidence = "medium"

    citation_bundle = _build_citation_bundle(
        item=item,
        narrative_matches=narrative_matches,
        supplementary_snippets=supplementary_snippets,
        policy=policy,
    )
    scope_classification = (
        str(narrative_matches[0].get("scope_classification", "unclassified"))
        if narrative_matches
        else "unclassified"
    )

    return {
        "metric": metric,
        "region": region,
        "card_type": "variance",
        "challenge_question": challenge_question,
        "prepared_answer": prepared_answer,
        "why_now": why_now,
        "basis_summary": basis_summary,
        "narrative_evidence_refs": sorted(set(narrative_refs))[:4],
        "supplementary_evidence_refs": sorted(set(supplementary_refs))[:6],
        "narrative_block_classes": sorted(set(narrative_classes))[:4],
        "scope_classification": scope_classification,
        "citation_bundle": citation_bundle,
        "confidence": confidence,
        "verify_next": verify_next,
    }


def _build_le_watchout_card(
    *,
    le_change_flags: list[dict[str, Any]],
    ranked_candidates: list[dict[str, Any]],
    supplementary_snippets: dict[str, list[dict[str, Any]]],
    policy: dict[str, Any],
) -> dict[str, Any]:
    sorted_flags = sorted(
        le_change_flags,
        key=lambda item: (
            1 if item.get("status") == "missing_current_period_le" else 0,
            abs(float(item.get("le_change") or 0.0)),
        ),
        reverse=True,
    )
    chosen = sorted_flags[0] if sorted_flags else {}
    metric = str(chosen.get("metric", ranked_candidates[0]["metric"] if ranked_candidates else "Total Sales"))
    candidate = next((item for item in ranked_candidates if item.get("metric") == metric), ranked_candidates[0] if ranked_candidates else {})
    unit = str(candidate.get("unit", "mm"))
    basis_summary = _format_basis_summary(candidate, unit) if candidate else {
        "vs_budget": "n/a",
        "vs_le": "LE not populated",
        "mom": "n/a",
        "qoq": "n/a",
        "yoy": "n/a",
    }
    narrative_matches = candidate.get("narrative_matches", []) if candidate else []
    narrative_refs = [str(match.get("evidence_ref", "")) for match in narrative_matches if match.get("evidence_ref")]
    narrative_classes = [str(match.get("block_class", "")) for match in narrative_matches if match.get("block_class")]
    supplementary_refs = [str(snip.get("evidence_ref", "")) for snip in supplementary_snippets.get(metric, []) if snip.get("evidence_ref")]
    region = str(candidate.get("region_hint", "C&US")) if candidate else "C&US"

    if chosen:
        if chosen.get("status") == "missing_current_period_le":
            challenge = f"{region} {metric} LE watchout: current period LE is not populated. How do we justify actual versus budget without LE anchor?"
            prepared = (
                f"LE not populated for {metric} in current period while prior period carried "
                f"{chosen.get('previous_le')}. Treat this as LE completeness risk, not automatic unfavorable variance."
            )
        else:
            le_change = chosen.get("le_change")
            le_delta = _format_signed_pp(le_change) if metric in {"SSS%", "SST%"} else _format_signed_mm(le_change)
            le_change_label = str(policy.get("semantics", {}).get("le_change_label", "vs prior LE"))
            challenge = (
                f"{region} {metric} LE watchout: LE shifted {le_delta} {le_change_label}. "
                "What changed in assumptions and why now?"
            )
            prepared = (
                f"LE changed {le_delta} for {metric}. Validate whether movement is structural (run-rate) or one-time timing."
            )
    else:
        challenge = f"{region} LE watchout: no material LE shifts detected. Which assumptions are most at risk of moving in next close?"
        prepared = "No material LE shifts detected in extracted metrics; maintain proactive watch on pricing, labor, and traffic assumptions."

    citation_bundle = _build_citation_bundle(
        item=candidate if candidate else {"metric": metric},
        narrative_matches=narrative_matches[:2],
        supplementary_snippets=supplementary_snippets.get(metric, [])[:2],
        policy=policy,
    )
    scope_classification = (
        str(narrative_matches[0].get("scope_classification", "unclassified"))
        if narrative_matches
        else "unclassified"
    )

    return {
        "metric": metric,
        "region": region,
        "card_type": "le_watchout",
        "challenge_question": challenge,
        "prepared_answer": prepared,
        "why_now": "LE deltas can change executive challenge framing even when bridge headlines are stable.",
        "basis_summary": basis_summary,
        "narrative_evidence_refs": sorted(set(narrative_refs))[:3],
        "supplementary_evidence_refs": sorted(set(supplementary_refs))[:4],
        "narrative_block_classes": sorted(set(narrative_classes))[:4],
        "scope_classification": scope_classification,
        "citation_bundle": citation_bundle,
        "confidence": "medium" if chosen else "low",
        "verify_next": "Confirm LE assumptions and ownership in latest workbook controls before leadership review.",
    }


def _card_to_hot_question(card: dict[str, Any]) -> dict[str, str]:
    return {
        "question": str(card.get("challenge_question", "")),
        "answer": f"{card.get('prepared_answer', '')} Verify next: {card.get('verify_next', '')}",
    }


def _build_variance_hot_questions(pack_dir: Path, policy: dict[str, Any]) -> dict[str, Any]:
    period = _infer_period_from_pack_dir(pack_dir)
    pack_type = _infer_pack_type_from_pack_dir(pack_dir)
    policy_version = str(policy.get("policy_version", "unknown"))
    scope_filters_applied: list[dict[str, Any]] = []
    evidence_gap_registry: list[dict[str, Any]] = []
    empty_payload = {
        "policy_version": policy_version,
        "term_guard_hits": [],
        "scope_filters_applied": scope_filters_applied,
        "evidence_gap_registry": evidence_gap_registry,
        "challenge_cards": [],
        "anticipated_hot_questions": [],
        "variance_question_candidates": [],
        "supplementary_metric_snippets": {},
        "le_change_flags": [],
        "supplementary_evidence_refs": [],
        "narrative_signal_summary": {},
        "le_completeness_watchouts": [],
        "quality_gate": _build_quality_gate([]),
    }
    if pack_type not in {"preview", "close"}:
        return empty_payload

    current_snapshot, _ = _load_metric_snapshot(pack_dir)
    if not current_snapshot:
        evidence_gap_registry.append(
            {
                "scope": "snapshot",
                "status": "not_helpful",
                "reason": "No current metric snapshot available from normalized workbook extracts.",
            }
        )
        return empty_payload

    repo_root = _repo_root_from_pack_dir(pack_dir)
    previous_snapshot: dict[str, dict[str, Any]] = {}
    qoq_snapshot: dict[str, dict[str, Any]] = {}
    prev_period = _shift_period(period, -1)
    qoq_period = _shift_period(period, -3)

    if repo_root:
        if prev_period:
            prev_pack = repo_root / "data" / "normalized" / prev_period / pack_type
            if prev_pack.exists():
                previous_snapshot, _ = _load_metric_snapshot(prev_pack)
        if qoq_period:
            qoq_pack = repo_root / "data" / "normalized" / qoq_period / pack_type
            if qoq_pack.exists():
                qoq_snapshot, _ = _load_metric_snapshot(qoq_pack)

    sales_now = current_snapshot.get("Total Sales", {}).get("actual")
    sales_prev = previous_snapshot.get("Total Sales", {}).get("actual")

    candidates: list[dict[str, Any]] = []
    for metric_name, spec in VARIANCE_METRIC_SPECS.items():
        current = current_snapshot.get(metric_name)
        if not current:
            continue
        actual = current.get("actual")
        if actual is None:
            continue
        budget = current.get("budget")
        le = current.get("le")
        prior_year = current.get("prior_year")
        prev_actual = previous_snapshot.get(metric_name, {}).get("actual")
        qoq_actual = qoq_snapshot.get(metric_name, {}).get("actual")
        prev_le = previous_snapshot.get(metric_name, {}).get("le")

        mom_delta = (actual - prev_actual) if prev_actual is not None else None
        qoq_delta = (actual - qoq_actual) if qoq_actual is not None else None
        vs_budget = (actual - budget) if budget is not None else None
        vs_le = (actual - le) if le is not None and abs(float(le)) > 1e-9 else None
        vs_py = (actual - prior_year) if prior_year is not None else None

        if spec["unit"] == "pct":
            severity = max(
                abs((mom_delta or 0.0) * 100),
                abs((qoq_delta or 0.0) * 100),
                abs((vs_budget or 0.0) * 100) * 1.4,
                abs((vs_le or 0.0) * 100) * 1.3,
                abs((vs_py or 0.0) * 100) * 0.8,
            )
        else:
            severity = max(
                abs(mom_delta or 0.0),
                abs(qoq_delta or 0.0),
                abs(vs_budget or 0.0) * 1.4,
                abs(vs_le or 0.0) * 1.3,
                abs(vs_py or 0.0) * 0.8,
            )

        margin_delta_pp: float | None = None
        if metric_name in {"AOI", "Total EBITDA"} and sales_now not in (None, 0) and sales_prev not in (None, 0):
            margin_now = actual / float(sales_now)
            margin_prev = prev_actual / float(sales_prev) if prev_actual is not None else None
            if margin_prev is not None:
                margin_delta_pp = (margin_now - margin_prev) * 100.0

        candidates.append(
            {
                "metric": metric_name,
                "unit": spec["unit"],
                "driver_focus": spec["driver_focus"],
                "row_number": current.get("row_number"),
                "evidence_ref": current.get("evidence_ref"),
                "actual": actual,
                "budget": budget,
                "le": le,
                "prior_year": prior_year,
                "prev_actual": prev_actual,
                "qoq_actual": qoq_actual,
                "previous_le": prev_le,
                "mom_delta": mom_delta,
                "qoq_delta": qoq_delta,
                "vs_budget": vs_budget,
                "vs_le": vs_le,
                "vs_py": vs_py,
                "mom_pct": _safe_percent_delta(actual, prev_actual) if spec["unit"] == "mm" else None,
                "qoq_pct": _safe_percent_delta(actual, qoq_actual) if spec["unit"] == "mm" else None,
                "vs_budget_pct": _safe_percent_delta(actual, budget) if spec["unit"] == "mm" else None,
                "vs_le_pct": _safe_percent_delta(actual, le) if spec["unit"] == "mm" else None,
                "vs_py_pct": _safe_percent_delta(actual, prior_year) if spec["unit"] == "mm" else None,
                "margin_delta_pp": margin_delta_pp,
                "le_change_vs_prior_month": (le - prev_le)
                if le is not None and prev_le is not None
                else None,
                "severity_score": round(severity, 3),
            }
        )

    narrative_blocks, narrative_summary = _collect_narrative_blocks(pack_dir)
    for candidate in candidates:
        metric = str(candidate.get("metric", ""))
        narrative_matches = _metric_narrative_matches(metric, narrative_blocks, policy)[:6]
        narrative_matches = _apply_scope_filters(
            narrative_matches,
            metric=metric,
            policy=policy,
            scope_filters_applied=scope_filters_applied,
        )[:4]
        candidate["narrative_matches"] = narrative_matches
        candidate["region_hint"] = str(narrative_matches[0].get("region", "C&US")) if narrative_matches else "C&US"
        candidate["scope_classification"] = (
            str(narrative_matches[0].get("scope_classification", "unclassified"))
            if narrative_matches
            else "unclassified"
        )
        if not narrative_matches:
            evidence_gap_registry.append(
                {
                    "scope": "narrative_match",
                    "metric": metric,
                    "status": "not_helpful",
                    "reason": "No restaurant-first narrative evidence remained after scope filtering.",
                }
            )
        narrative_bonus = float(narrative_matches[0].get("score", 0.0)) if narrative_matches else -4.0
        candidate["ranking_score"] = round(float(candidate.get("severity_score", 0.0)) + max(-6.0, narrative_bonus), 3)

    ranked = sorted(candidates, key=lambda item: float(item.get("ranking_score", 0.0)), reverse=True)
    supplementary_snippets, supplementary_evidence_refs, supplementary_registry = _collect_supplementary_evidence(
        pack_dir,
        ranked,
        policy,
    )
    evidence_gap_registry.extend(supplementary_registry)
    le_change_flags = _compute_le_change_flags(current_snapshot, previous_snapshot, policy)
    card_rules = policy.get("card_rules", {})
    min_cards = int(card_rules.get("min_cards", 2))
    target_cards = int(card_rules.get("target_cards", 5))
    max_cards = int(card_rules.get("max_cards", 5))

    def _candidate_supportable(candidate: dict[str, Any]) -> bool:
        basis_summary = _format_basis_summary(candidate, str(candidate.get("unit", "mm")))
        return bool(candidate.get("narrative_matches")) and _ensure_basis_presence(basis_summary)

    supportable_candidates = [candidate for candidate in ranked if _candidate_supportable(candidate)]
    if len(supportable_candidates) < min_cards:
        evidence_gap_registry.append(
            {
                "scope": "card_sufficiency",
                "status": "not_helpful",
                "reason": (
                    f"Only {len(supportable_candidates)} evidence-backed cards supportable; "
                    f"minimum required is {min_cards}."
                ),
            }
        )

    selected_items: list[tuple[str, dict[str, Any]]] = []
    used_metrics: set[str] = set()
    required_mix = [("C&US", 2), ("Canada", 1), ("US", 1)]
    for region, required_count in required_mix:
        count = 0
        region_candidates = [
            item
            for item in supportable_candidates
            if item.get("region_hint") == region and item.get("metric") not in used_metrics
        ]
        fallback_candidates = [item for item in supportable_candidates if item.get("metric") not in used_metrics]
        for candidate in [*region_candidates, *fallback_candidates]:
            metric = str(candidate.get("metric", ""))
            if not metric or metric in used_metrics:
                continue
            selected_items.append((region, candidate))
            used_metrics.add(metric)
            count += 1
            if count >= required_count:
                break

    challenge_cards: list[dict[str, Any]] = []
    for region, candidate in selected_items[:target_cards]:
        metric = str(candidate.get("metric", ""))
        challenge_cards.append(
            _build_challenge_card(
                item=candidate,
                period=period,
                region=region,
                narrative_matches=(
                    [match for match in candidate.get("narrative_matches", []) if match.get("block_class") == "narrative"][:2]
                    or candidate.get("narrative_matches", [])[:2]
                ),
                supplementary_snippets=supplementary_snippets.get(metric, [])[:2],
                policy=policy,
            )
        )

    if len(challenge_cards) < target_cards:
        for candidate in supportable_candidates:
            metric = str(candidate.get("metric", ""))
            if metric in used_metrics:
                continue
            challenge_cards.append(
                _build_challenge_card(
                    item=candidate,
                    period=period,
                    region=str(candidate.get("region_hint", "C&US")),
                    narrative_matches=candidate.get("narrative_matches", [])[:2],
                    supplementary_snippets=supplementary_snippets.get(metric, [])[:2],
                    policy=policy,
                )
            )
            used_metrics.add(metric)
            if len(challenge_cards) >= target_cards:
                break

    watchout_rules = policy.get("le_watchout_rules", {})
    completeness_flags = [item for item in le_change_flags if item.get("status") == "missing_current_period_le"]
    material_flags = [item for item in le_change_flags if item.get("status") != "missing_current_period_le"]
    include_watchout = bool(watchout_rules.get("enabled", True)) and (
        (bool(watchout_rules.get("include_on_material_shift", True)) and bool(material_flags))
        or (bool(watchout_rules.get("include_on_completeness_gap", True)) and bool(completeness_flags))
    )
    if include_watchout:
        watchout_card = _build_le_watchout_card(
            le_change_flags=le_change_flags,
            ranked_candidates=ranked,
            supplementary_snippets=supplementary_snippets,
            policy=policy,
        )
        if len(challenge_cards) >= max_cards:
            challenge_cards = challenge_cards[: max(0, max_cards - 1)]
        challenge_cards.append(watchout_card)
    challenge_cards = challenge_cards[:max_cards]

    if len([card for card in challenge_cards if card.get("card_type") != "le_watchout"]) < min_cards:
        evidence_gap_registry.append(
            {
                "scope": "card_output",
                "status": "not_helpful",
                "reason": (
                    f"Returned fewer than minimum required evidence-backed cards ({min_cards}); "
                    "output intentionally not padded."
                ),
            }
        )

    le_completeness_watchouts = [
        {
            "metric": str(item.get("metric", "")),
            "status": str(item.get("status", "")),
            "message": f"{item.get('metric')} LE not populated in current period; verify template load and forecast governance.",
        }
        for item in le_change_flags
        if item.get("status") == "missing_current_period_le"
    ]
    challenge_cards, term_guard_hits = _apply_term_guard_to_cards(challenge_cards, policy=policy)
    anticipated_hot_questions = [_card_to_hot_question(card) for card in challenge_cards]
    quality_gate = _build_quality_gate(challenge_cards)

    return {
        "policy_version": policy_version,
        "term_guard_hits": term_guard_hits,
        "scope_filters_applied": scope_filters_applied,
        "evidence_gap_registry": evidence_gap_registry,
        "challenge_cards": challenge_cards,
        "anticipated_hot_questions": anticipated_hot_questions,
        "variance_question_candidates": ranked[:10],
        "supplementary_metric_snippets": supplementary_snippets,
        "le_change_flags": le_change_flags,
        "supplementary_evidence_refs": supplementary_evidence_refs[:50],
        "narrative_signal_summary": narrative_summary,
        "le_completeness_watchouts": le_completeness_watchouts,
        "quality_gate": quality_gate,
    }


def run_hot_questions(
    pack_dir: Path,
    question: str | None = None,
    *,
    scoring_config: dict[str, Any] | None = None,
    policy_config: dict[str, Any] | None = None,
    month_override: dict[str, Any] | None = None,
    historical_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    config = load_hotq_scoring_config() if scoring_config is None else scoring_config
    policy = load_hotq_policy() if policy_config is None else policy_config
    weights = config["weights"]
    penalties = config.get("penalties", {})
    pack_type = _infer_pack_type_from_pack_dir(pack_dir)

    slide_files = _iter_slide_files(pack_dir)
    workbook_meta_files = _iter_workbook_meta(pack_dir)
    pack_summary = _read_pack_summary(pack_dir)

    total_slides = len(slide_files)
    total_workbooks = len(workbook_meta_files)
    notes_missing = 0
    favorable_count = 0
    unfavorable_count = 0
    fx_mentions = 0
    inflation_mentions = 0
    risk_mentions = 0
    performance_keyword_hits = 0
    sign_mismatch_signals = 0
    evidence: list[str] = []

    for slide_path in slide_files:
        slide = read_json(slide_path)
        title = slide.get("title", "")
        body = " ".join(slide.get("body", []))
        note_text = slide.get("note_text", "")
        combined_text = f"{title} {body} {note_text}".lower()
        if not note_text:
            notes_missing += 1
        if "favorable" in combined_text:
            favorable_count += 1
        if "unfavorable" in combined_text:
            unfavorable_count += 1
        if "fx" in combined_text or "foreign exchange" in combined_text:
            fx_mentions += 1
        if "inflation" in combined_text:
            inflation_mentions += 1
        for token in ["budget", "le", "plan", "py", "yoy", "wow", "sss", "sst"]:
            if token in combined_text:
                performance_keyword_hits += 1
        for token in ["risk", "pressure", "headwind", "volatility", "challenge", "competitive"]:
            if token in combined_text:
                risk_mentions += 1
        if "favorable" in combined_text and "(" in combined_text and ")" in combined_text:
            sign_mismatch_signals += 1
        if len(evidence) < 8:
            evidence.append(str(slide_path.relative_to(pack_dir).as_posix()))

    external_link_count = 0
    formula_cells = 0
    external_formula_cells = 0
    bridge_sheet_count = 0
    for meta_path in workbook_meta_files:
        workbook_dir = meta_path.parent
        meta = read_json(meta_path)
        lineage_flags = read_json(workbook_dir / "lineage_flags.json")
        external_link_count += int(lineage_flags.get("external_link_count", 0))
        formula_cells += int(lineage_flags.get("formula_cells_total", 0))
        external_formula_cells += int(lineage_flags.get("external_formula_cells_total", 0))
        bridge_sheet_count += sum(
            1 for sheet in meta.get("sheets", []) if "bridge" in str(sheet.get("sheet_name", "")).lower()
        )
        if len(evidence) < 12:
            evidence.append(str(meta_path.relative_to(pack_dir).as_posix()))

    source_mode = str(pack_summary.get("source_mode", "unknown"))
    lineage_degraded = bool(pack_summary.get("lineage_degraded", False))

    notes_missing_ratio = (notes_missing / total_slides) if total_slides else 1.0

    dimension_scores: list[dict[str, Any]] = []

    pnl_raw = (
        52
        + min(24.0, performance_keyword_hits * 1.5)
        + (favorable_count - unfavorable_count) * 2.5
        - max(0.0, unfavorable_count - favorable_count) * 1.5
    )
    pnl_score = _clamp_score(pnl_raw)
    dimension_scores.append(
        {
            "dimension": "pnl_delivery",
            "weight": weights.get("pnl_delivery", 30),
            "score": round(pnl_score, 2),
            "drivers": [
                f"performance_keyword_hits={performance_keyword_hits}",
                f"favorable={favorable_count}",
                f"unfavorable={unfavorable_count}",
            ],
        }
    )

    variance_raw = (
        30
        + min(35.0, formula_cells / 1200.0 * 35.0)
        + (20.0 if bridge_sheet_count > 0 else -10.0)
        + min(20.0, external_formula_cells / 250.0 * 20.0)
    )
    variance_score = _clamp_score(variance_raw)
    dimension_scores.append(
        {
            "dimension": "variance_explainability",
            "weight": weights.get("variance_explainability", 25),
            "score": round(variance_score, 2),
            "drivers": [
                f"formula_cells={formula_cells}",
                f"bridge_sheet_count={bridge_sheet_count}",
                f"external_formula_cells={external_formula_cells}",
            ],
        }
    )

    forecast_raw = (
        68
        - notes_missing_ratio * 26.0
        - (6.0 if risk_mentions > 25 else 0.0)
        + min(12.0, risk_mentions * 0.5)
        + min(5.0, inflation_mentions * 0.5)
    )
    forecast_score = _clamp_score(forecast_raw)
    dimension_scores.append(
        {
            "dimension": "forecast_reliability",
            "weight": weights.get("forecast_reliability", 20),
            "score": round(forecast_score, 2),
            "drivers": [
                f"notes_missing_ratio={notes_missing_ratio:.2f}",
                f"risk_mentions={risk_mentions}",
                f"inflation_mentions={inflation_mentions}",
            ],
        }
    )

    narrative_raw = 82 - notes_missing_ratio * 30.0 - sign_mismatch_signals * 8.0
    narrative_score = _clamp_score(narrative_raw)
    dimension_scores.append(
        {
            "dimension": "narrative_integrity",
            "weight": weights.get("narrative_integrity", 15),
            "score": round(narrative_score, 2),
            "drivers": [
                f"notes_missing={notes_missing}",
                f"sign_mismatch_signals={sign_mismatch_signals}",
            ],
        }
    )

    data_confidence_raw = (
        45
        + min(20.0, external_link_count * 1.5)
        + min(20.0, external_formula_cells / 250.0 * 20.0)
        + (10.0 if not lineage_degraded else -18.0)
    )
    data_confidence_score = _clamp_score(data_confidence_raw)
    dimension_scores.append(
        {
            "dimension": "data_confidence",
            "weight": weights.get("data_confidence", 10),
            "score": round(data_confidence_score, 2),
            "drivers": [
                f"external_link_count={external_link_count}",
                f"external_formula_cells={external_formula_cells}",
                f"lineage_degraded={lineage_degraded}",
            ],
        }
    )

    override_notes: list[str] = []
    if month_override:
        for adjustment in month_override.get("score_adjustments", []):
            target = str(adjustment.get("dimension", ""))
            delta = float(adjustment.get("delta", 0))
            reason = str(adjustment.get("reason", "manual adjustment"))
            for item in dimension_scores:
                if item["dimension"] == target:
                    item["score"] = round(_clamp_score(float(item["score"]) + delta), 2)
                    override_notes.append(f"{target}:{delta:+.1f} ({reason})")
                    break

    historical_notes: list[str] = []
    if historical_context:
        calibrated_deltas = (
            historical_context.get("calibrated_deltas", {}).get(pack_type, {})
            if isinstance(historical_context.get("calibrated_deltas"), dict)
            else {}
        )
        for target, raw_delta in calibrated_deltas.items():
            delta = float(raw_delta)
            for item in dimension_scores:
                if item["dimension"] == target:
                    item["score"] = round(_clamp_score(float(item["score"]) + delta), 2)
                    historical_notes.append(f"{target}:{delta:+.2f} (historical calibration)")
                    break

    weighted_points = 0.0
    for item in dimension_scores:
        weight = float(item["weight"])
        item["weighted_points"] = round(weight * float(item["score"]) / 100.0, 2)
        weighted_points += item["weighted_points"]

    penalty_total = 0.0
    if lineage_degraded:
        penalty_total += float(penalties.get("lineage_degraded", 0))
    if notes_missing_ratio >= 0.60:
        penalty_total += float(penalties.get("notes_missing_ratio_high", 0))
    if total_slides == 0 or total_workbooks == 0:
        penalty_total += float(penalties.get("missing_pack_components", 0))

    global_delta = float((month_override or {}).get("global_delta", 0))
    score_total = _clamp_score(weighted_points - penalty_total + global_delta)
    score_band = _score_band(score_total, config.get("thresholds", {}))

    confidence = "high"
    confidence_reason = "Pack contains both slide and workbook evidence with lineage coverage."
    clarifier = ""
    if total_slides == 0 or total_workbooks == 0:
        confidence = "low"
        confidence_reason = "Missing deck or workbook coverage in normalized pack."
        clarifier = (
            "I do not have a complete normalized pack yet. Please confirm that both deck and workbook files "
            "were ingested for this period."
        )
    elif external_link_count == 0 and external_formula_cells == 0:
        confidence = "medium"
        confidence_reason = "No external-link lineage detected in current extracted workbook content."
        clarifier = (
            "Lineage signals are limited (no external links/formula lineage found). "
            "Do you want me to treat this as offline-values-only for root-cause confidence?"
        )

    if month_override and month_override.get("force_clarifier"):
        clarifier = str(month_override["force_clarifier"])

    trailing_context: list[str] = []
    baseline_note = ""
    if historical_context:
        score_baseline = (
            historical_context.get("score_baselines", {}).get(pack_type, {})
            if isinstance(historical_context.get("score_baselines"), dict)
            else {}
        )
        baseline_mean = score_baseline.get("mean_score_total")
        if isinstance(baseline_mean, (int, float)):
            delta_vs_baseline = round(score_total - float(baseline_mean), 2)
            baseline_note = (
                f"Historical baseline ({pack_type}): {float(baseline_mean):.1f} "
                f"(delta {delta_vs_baseline:+.2f})."
            )
        trailing_map = historical_context.get("trailing_period_context", {})
        if isinstance(trailing_map, dict):
            trailing_values = trailing_map.get(pack_type, [])
            if isinstance(trailing_values, list):
                trailing_context = [str(item) for item in trailing_values[:4]]

    summary_bullets = [
        f"Scorecard: {score_total:.1f} ({score_band}) | confidence={confidence}.",
        f"Coverage: {total_slides} normalized slides across {total_workbooks} workbook extracts.",
        f"Commentary tone: favorable={favorable_count}, unfavorable={unfavorable_count}, risk_mentions={risk_mentions}.",
        f"Lineage density: formula_cells={formula_cells}, external_links={external_link_count}, external_formula_cells={external_formula_cells}.",
    ]
    if pack_type == "preview" and bool(policy.get("semantics", {}).get("preview_equals_le", True)):
        summary_bullets.append(
            "Semantic guard: preview is treated as current LE; use vs Budget and vs prior LE for LE movement."
        )
    if baseline_note:
        summary_bullets.append(baseline_note)
    if trailing_context:
        summary_bullets.append(f"Trailing context: {' | '.join(trailing_context)}")
    if question:
        summary_bullets.insert(0, f"Question focus: {question}")

    compact_table = [
        {
            "metric": "Score Band",
            "observation": f"{score_total:.1f} ({score_band})",
            "implication": "Use band to triage depth of executive follow-up and forecast challenge.",
        },
        {
            "metric": "Narrative Coverage",
            "observation": f"{notes_missing}/{total_slides} slides missing note text",
            "implication": "Lower note coverage increases risk of unsupported commentary in leadership review.",
        },
        {
            "metric": "Lineage Confidence",
            "observation": f"{external_formula_cells} external-reference formula cells | source_mode={source_mode}",
            "implication": "Low lineage can reduce attribution confidence for variance root causes.",
        },
    ]

    risks = [
        "Potential mismatch risk where narrative does not include explicit numeric support.",
        "Potential stale assumptions if prior-month commentary was rolled forward without updates.",
    ]
    opportunities = [
        "Use slide-level numeric mention mapping to tighten commentary-to-chart consistency.",
        "Promote recurring variance drivers into a standing monthly risk/opportunity template.",
    ]
    actions = [
        "Run deck proofing before final executive circulation.",
        "Run variance watch on bridge sheets to validate driver math and period consistency.",
    ]
    if score_band == "Red":
        actions.insert(0, "Initiate targeted forecast challenge review on top variance drivers this cycle.")
    if score_band == "Yellow":
        actions.insert(0, "Prioritize top 3 medium/high-risk issues before executive sign-off.")

    hotq_bundle = _build_variance_hot_questions(pack_dir, policy)
    challenge_cards = hotq_bundle.get("challenge_cards", [])
    anticipated_hot_questions = hotq_bundle.get("anticipated_hot_questions", [])
    variance_question_candidates = hotq_bundle.get("variance_question_candidates", [])
    supplementary_metric_snippets = hotq_bundle.get("supplementary_metric_snippets", {})
    le_change_flags = hotq_bundle.get("le_change_flags", [])
    supplementary_evidence_refs = hotq_bundle.get("supplementary_evidence_refs", [])
    narrative_signal_summary = hotq_bundle.get("narrative_signal_summary", {})
    le_completeness_watchouts = hotq_bundle.get("le_completeness_watchouts", [])
    quality_gate = hotq_bundle.get("quality_gate", _build_quality_gate(challenge_cards))
    policy_version = str(hotq_bundle.get("policy_version", policy.get("policy_version", "unknown")))
    term_guard_hits = hotq_bundle.get("term_guard_hits", [])
    scope_filters_applied = hotq_bundle.get("scope_filters_applied", [])
    evidence_gap_registry = hotq_bundle.get("evidence_gap_registry", [])

    if variance_question_candidates:
        top = variance_question_candidates[0]
        top_metric = str(top.get("metric", "metric"))
        unit = str(top.get("unit", "mm"))
        summary_bullets.append(
            f"Top variance focus ({top_metric}): {_basis_phrase(top, basis='vs_budget', unit=unit)} | "
            f"{_basis_phrase(top, basis='vs_le', unit=unit)}."
        )
    if challenge_cards:
        top_card = challenge_cards[0]
        summary_bullets.append(
            f"Top challenge card: {top_card.get('region')} {top_card.get('metric')} | "
            f"{top_card.get('basis_summary', {}).get('vs_budget', 'n/a')} vs Budget, "
            f"{top_card.get('basis_summary', {}).get('vs_le', 'n/a')} vs LE."
        )
    for ref in supplementary_evidence_refs:
        if ref not in evidence and len(evidence) < 24:
            evidence.append(ref)
    for card in challenge_cards:
        for ref in card.get("narrative_evidence_refs", []):
            if ref and ref not in evidence and len(evidence) < 24:
                evidence.append(ref)
    if le_change_flags:
        top_le = sorted(le_change_flags, key=lambda item: abs(float(item.get("le_change") or 0.0)), reverse=True)[:2]
        le_notes = []
        for item in top_le:
            metric = str(item.get("metric", "metric"))
            if item.get("status") == "missing_current_period_le":
                le_notes.append(f"{metric} LE is missing/zero in current period; prior month carried a value.")
            elif metric in {"SSS%", "SST%"}:
                le_notes.append(f"{metric} LE moved {_format_signed_pp(item.get('le_change'))} vs prior month.")
            else:
                le_notes.append(f"{metric} LE moved {_format_signed_mm(item.get('le_change'))} vs prior month.")
        if le_notes:
            summary_bullets.append("LE shifts: " + " | ".join(le_notes))
    if le_completeness_watchouts:
        summary_bullets.append("LE completeness watchout: " + le_completeness_watchouts[0].get("message", ""))
    insufficiency_notice = ""
    if evidence_gap_registry:
        insufficiency = next(
            (
                item
                for item in evidence_gap_registry
                if str(item.get("scope", "")).startswith("card_") and item.get("status") == "not_helpful"
            ),
            None,
        )
        if insufficiency:
            insufficiency_notice = str(insufficiency.get("reason", "")).strip()
            summary_bullets.append(f"Evidence insufficiency: {insufficiency_notice}")

    if quality_gate.get("status") == "downgraded_narrative_gap":
        summary_bullets.append("Nuance gate: downgraded due to weak narrative evidence on at least one card.")
        if confidence == "high":
            confidence = "medium"
            confidence_reason = (
                "Narrative evidence coverage is partial; deterministic variance cards generated with downgrade warning."
            )
    elif quality_gate.get("status") == "fail":
        summary_bullets.append("Nuance gate: failed (insufficient narrative support or basis linkage).")
        confidence = "low"
        confidence_reason = "Challenge-card quality gate failed due to missing required narrative or basis evidence."

    if not anticipated_hot_questions:
        anticipated_hot_questions = [
            {
                "question": "What are the most likely variance challenges I should expect in leadership review?",
                "answer": (
                    f"Expect challenge on driver clarity and variance explainability. Current band is {score_band} "
                    f"with variance explainability score {variance_score:.1f}, bridge_sheet_count={bridge_sheet_count}, "
                    f"and formula_cells={formula_cells}."
                ),
            },
            {
                "question": "How confident are we in the numbers behind the variance narrative?",
                "answer": (
                    f"Data confidence is {data_confidence_score:.1f} with source_mode={source_mode}, "
                    f"external_formula_cells={external_formula_cells}, and lineage_degraded={lineage_degraded}."
                ),
            },
            {
                "question": "Where is the commentary most exposed to executive pushback?",
                "answer": (
                    f"Narrative integrity is {narrative_score:.1f}; {notes_missing}/{total_slides} slides are missing notes, "
                    f"and sign_mismatch_signals={sign_mismatch_signals}."
                ),
            },
        ]

    if not challenge_cards and not bool(policy.get("card_rules", {}).get("do_not_pad_placeholders", True)):
        challenge_cards = [
            {
                "metric": "Portfolio",
                "region": "C&US",
                "card_type": "variance",
                "challenge_question": anticipated_hot_questions[0]["question"],
                "prepared_answer": anticipated_hot_questions[0]["answer"],
                "why_now": "Narrative cards were unavailable from extracted evidence.",
                "basis_summary": {"vs_budget": "n/a", "vs_le": "LE not populated", "mom": "n/a", "qoq": "n/a", "yoy": "n/a"},
                "narrative_evidence_refs": [],
                "supplementary_evidence_refs": supplementary_evidence_refs[:2],
                "narrative_block_classes": [],
                "scope_classification": "unclassified",
                "citation_bundle": [],
                "confidence": "low",
                "verify_next": "Re-tokenize close pack and confirm slide text blocks are available.",
            }
        ]
        quality_gate = _build_quality_gate(challenge_cards)

    follow_up_prompt = "Is there any specific questions you'd like help coming up with an answer for?"

    payload = {
        "generated_at": utc_now_iso(),
        "period": _infer_period_from_pack_dir(pack_dir),
        "confidence": confidence,
        "confidence_reason": confidence_reason,
        "score_total": round(score_total, 2),
        "score_band": score_band,
        "dimension_scores": dimension_scores,
        "summary_bullets": summary_bullets[:6],
        "compact_table": compact_table,
        "risks": (month_override or {}).get("risk_overrides", risks),
        "opportunities": (month_override or {}).get("opportunity_overrides", opportunities),
        "actions": (month_override or {}).get("action_overrides", actions),
        "clarifying_question": clarifier,
        "evidence_refs": evidence,
        "source_mode": source_mode,
        "lineage_degraded": lineage_degraded,
        "applied_penalties": round(penalty_total, 2),
        "override_notes": override_notes,
        "historical_notes": historical_notes,
        "historical_context_applied": bool(historical_context),
        "policy_version": policy_version,
        "term_guard_hits": term_guard_hits,
        "scope_filters_applied": scope_filters_applied,
        "evidence_gap_registry": evidence_gap_registry,
        "insufficiency_notice": insufficiency_notice,
        "hot_question_prompt_version": VARIANCE_QUESTION_PROMPT_VERSION,
        "hot_question_prompt": VARIANCE_QUESTION_PROMPT,
        "variance_question_candidates": variance_question_candidates[:10],
        "supplementary_metric_snippets": supplementary_metric_snippets,
        "le_change_flags": le_change_flags,
        "supplementary_evidence_refs": supplementary_evidence_refs[:50],
        "challenge_cards": challenge_cards[: int(policy.get("card_rules", {}).get("max_cards", 5))],
        "narrative_signal_summary": narrative_signal_summary,
        "le_completeness_watchouts": le_completeness_watchouts,
        "quality_gate": quality_gate,
        "anticipated_hot_questions": anticipated_hot_questions,
        "follow_up_prompt": follow_up_prompt,
    }
    return payload


def run_deck_proofing(pack_dir: Path, prior_pack_dir: Path | None = None) -> dict[str, Any]:
    issues: list[Issue] = []
    current_slides = _iter_slide_files(pack_dir)
    prior_slide_map: dict[str, dict[str, Any]] = {}
    if prior_pack_dir and prior_pack_dir.exists():
        for slide_path in _iter_slide_files(prior_pack_dir):
            prior_slide_map[slide_path.name] = read_json(slide_path)

    confidentiality_hits = 0
    for slide_path in current_slides:
        slide = read_json(slide_path)
        title = slide.get("title", "").strip()
        body = " ".join(slide.get("body", []))
        note_text = slide.get("note_text", "").strip()
        combined = f"{title} {body} {note_text}".lower()
        slide_ref = str(slide_path.relative_to(pack_dir).as_posix())

        if not title or re.fullmatch(r"\d{1,3}", title):
            issues.append(
                Issue(
                    location=slide_ref,
                    issue_type="Missing",
                    description="Slide title appears missing or is only a page number.",
                    severity="Medium",
                    recommended_fix="Set an explicit descriptive title aligned to the financial message.",
                    evidence_refs=[slide_ref],
                )
            )

        if not note_text:
            issues.append(
                Issue(
                    location=slide_ref,
                    issue_type="Missing",
                    description="Slide has no note/commentary text.",
                    severity="Low",
                    recommended_fix="Add concise commentary to document driver context and caveats.",
                    evidence_refs=[slide_ref],
                )
            )

        if "favorable" in combined and "(" in combined and ")" in combined:
            issues.append(
                Issue(
                    location=slide_ref,
                    issue_type="Mismatch",
                    description=(
                        "Potential sign-language mismatch: commentary uses 'favorable' while parenthetical "
                        "negative formatting appears in same slide context."
                    ),
                    severity="Medium",
                    recommended_fix="Validate sign convention and revise wording or numeric formatting.",
                    evidence_refs=[slide_ref],
                )
            )

        if "confidential and proprietary information" in combined:
            confidentiality_hits += 1

        if "adjusted ebitda" in combined and "gaap" not in combined and "ifrs" not in combined:
            issues.append(
                Issue(
                    location=slide_ref,
                    issue_type="Consistency",
                    description="Adjusted EBITDA is referenced without explicit GAAP/IFRS labeling context.",
                    severity="Medium",
                    recommended_fix="Add GAAP/IFRS/non-GAAP labeling note per reporting policy.",
                    evidence_refs=[slide_ref],
                )
            )

        if prior_slide_map:
            prior = prior_slide_map.get(slide_path.name)
            if prior:
                prior_text = " ".join([prior.get("title", ""), " ".join(prior.get("body", []))]).strip()
                current_text = " ".join([title, body]).strip()
                if prior_text and prior_text == current_text:
                    issues.append(
                        Issue(
                            location=slide_ref,
                            issue_type="Outdated",
                            description="Slide commentary appears unchanged from prior pack.",
                            severity="Medium",
                            recommended_fix="Revalidate narrative against current-month metrics and update wording.",
                            evidence_refs=[slide_ref],
                        )
                    )

    if confidentiality_hits == 0:
        issues.append(
            Issue(
                location="deck_meta",
                issue_type="Missing",
                description="No confidentiality disclaimer text detected in slide content.",
                severity="Low",
                recommended_fix="Confirm deck template includes required confidentiality footer/disclaimer.",
                evidence_refs=["decks"],
            )
        )

    payload = {
        "generated_at": utc_now_iso(),
        "confidence": "high",
        "issue_count": len(issues),
        "issues": [issue.as_dict() for issue in issues],
        "evidence_refs": sorted({ref for issue in issues for ref in issue.evidence_refs})[:100],
    }
    return payload


def _read_sheet_strings(values_csv_path: Path, limit_rows: int = 180) -> list[str]:
    values: list[str] = []
    if not values_csv_path.exists():
        return values
    with values_csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        next(reader, None)
        for index, row in enumerate(reader, start=1):
            if index > limit_rows:
                break
            values.extend([cell for cell in row if cell and not re.fullmatch(r"-?\d+(?:\.\d+)?", cell)])
    return values


def run_variance_watch(pack_dir: Path) -> dict[str, Any]:
    issues: list[Issue] = []
    workbook_meta_files = _iter_workbook_meta(pack_dir)

    for meta_path in workbook_meta_files:
        meta = read_json(meta_path)
        workbook_dir = meta_path.parent
        for sheet in meta.get("sheets", []):
            sheet_name = sheet.get("sheet_name", "")
            location = f"{meta_path.parent.name}/{sheet_name}"

            formula_cells = int(sheet.get("formula_cells", 0))
            external_formula_cells = int(sheet.get("external_formula_cells", 0))
            values_csv_path = workbook_dir / sheet.get("values_csv", "")
            strings = [text.lower() for text in _read_sheet_strings(values_csv_path)]
            text_blob = " ".join(strings)

            if "bridge" in sheet_name.lower() and formula_cells == 0:
                issues.append(
                    Issue(
                        location=location,
                        issue_type="Mismatch",
                        description="Bridge sheet has no formulas in extracted range; bridge integrity cannot be validated.",
                        severity="High",
                        recommended_fix="Expand extraction range or confirm formula links are preserved in source workbook.",
                        evidence_refs=[str(values_csv_path.relative_to(pack_dir).as_posix())],
                    )
                )

            period_tokens = {"ytd": "YTD", "qtd": "QTD", "fy": "FY"}
            present_periods = [label for token, label in period_tokens.items() if token in text_blob]
            if len(present_periods) > 1:
                issues.append(
                    Issue(
                        location=location,
                        issue_type="Consistency",
                        description=f"Multiple period tokens detected on one sheet: {', '.join(present_periods)}.",
                        severity="Medium",
                        recommended_fix="Confirm period basis is consistent for each bridge/variance statement.",
                        evidence_refs=[str(values_csv_path.relative_to(pack_dir).as_posix())],
                    )
                )

            if "fx" in text_blob and "rate" not in text_blob:
                issues.append(
                    Issue(
                        location=location,
                        issue_type="Consistency",
                        description="FX references detected without explicit rate context.",
                        severity="Medium",
                        recommended_fix="Add plan-rate vs actual-rate support where FX variance is discussed.",
                        evidence_refs=[str(values_csv_path.relative_to(pack_dir).as_posix())],
                    )
                )

            for marker in ["tbu", "tbd", "placeholder", "xxx"]:
                if marker in text_blob:
                    issues.append(
                        Issue(
                            location=location,
                            issue_type="Missing",
                            description=f"Incomplete marker '{marker.upper()}' detected in sheet content.",
                            severity="High",
                            recommended_fix="Replace placeholder with final commentary or schedule values.",
                            evidence_refs=[str(values_csv_path.relative_to(pack_dir).as_posix())],
                        )
                    )
                    break

            if external_formula_cells == 0 and "check" in sheet_name.lower():
                issues.append(
                    Issue(
                        location=location,
                        issue_type="Clarity",
                        description=(
                            "Validation/check sheet has no external-reference formulas in extracted cells. "
                            "Lineage coverage may be incomplete."
                        ),
                        severity="Low",
                        recommended_fix="Confirm this check tab is expected to run offline or include lineage export.",
                        evidence_refs=[str(values_csv_path.relative_to(pack_dir).as_posix())],
                    )
                )

            formula_csv = workbook_dir / sheet.get("formula_cells_csv", "")
            if formula_csv.exists():
                with formula_csv.open("r", encoding="utf-8", newline="") as handle:
                    reader = csv.DictReader(handle)
                    has_driver_headers = False
                    for row in reader:
                        formula = (row.get("formula") or "").lower()
                        if all(token in formula for token in ["price", "volume", "mix"]):
                            has_driver_headers = True
                            break
                    if not has_driver_headers and "bridge" in sheet_name.lower():
                        issues.append(
                            Issue(
                                location=location,
                                issue_type="Clarity",
                                description=(
                                    "Bridge formulas do not clearly evidence canonical price/volume/mix driver structure."
                                ),
                                severity="Low",
                                recommended_fix="Verify driver decomposition aligns to reporting convention.",
                                evidence_refs=[str(formula_csv.relative_to(pack_dir).as_posix())],
                            )
                        )

    payload = {
        "generated_at": utc_now_iso(),
        "confidence": "high",
        "issue_count": len(issues),
        "issues": [issue.as_dict() for issue in issues],
        "evidence_refs": sorted({ref for issue in issues for ref in issue.evidence_refs})[:200],
    }
    return payload


def persist_analysis(payload: dict[str, Any], output_path: Path) -> None:
    write_json(output_path, payload)

