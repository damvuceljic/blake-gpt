from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from finance_copilot.common import (
    ensure_dir,
    is_external_formula,
    slugify,
    to_text,
    utc_now_iso,
    write_json,
)


def _extract_named_ranges(workbook: Any) -> dict[str, list[dict[str, str]]]:
    per_sheet: dict[str, list[dict[str, str]]] = defaultdict(list)
    for name in workbook.defined_names.keys():
        try:
            defined_name = workbook.defined_names[name]
        except Exception:
            continue
        try:
            destinations = list(defined_name.destinations)
        except Exception:
            destinations = []
        for sheet_name, reference in destinations:
            per_sheet[sheet_name].append({"name": str(name), "reference": str(reference)})
    return per_sheet


def _extract_external_links(workbook: Any) -> list[dict[str, str]]:
    links: list[dict[str, str]] = []
    for index, link in enumerate(getattr(workbook, "_external_links", []) or [], start=1):
        file_link = getattr(link, "file_link", None)
        target = getattr(file_link, "Target", "")
        rel_id = getattr(file_link, "id", "")
        links.append(
            {
                "index": str(index),
                "target": str(target or ""),
                "relationship_id": str(rel_id or ""),
            }
        )
    return links


def extract_workbook(
    *,
    input_path: Path,
    output_dir: Path,
    max_rows: int | None = None,
    max_cols: int | None = None,
) -> dict[str, Any]:
    ensure_dir(output_dir)
    sheets_dir = ensure_dir(output_dir / "sheets")

    workbook = load_workbook(input_path, read_only=True, data_only=False)
    named_ranges_by_sheet = _extract_named_ranges(workbook)
    external_links = _extract_external_links(workbook)

    workbook_meta: dict[str, Any] = {
        "source_file": input_path.name,
        "extracted_at": utc_now_iso(),
        "sheet_count": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames,
        "sheets": [],
    }

    formula_total = 0
    external_formula_total = 0

    for sheet in workbook.worksheets:
        sheet_slug = slugify(sheet.title)
        sheet_out = ensure_dir(sheets_dir / sheet_slug)
        used_rows = min(sheet.max_row or 0, max_rows) if max_rows else (sheet.max_row or 0)
        used_cols = min(sheet.max_column or 0, max_cols) if max_cols else (sheet.max_column or 0)

        values_path = sheet_out / "values.csv"
        formulas_path = sheet_out / "formula_cells.csv"
        named_ranges_path = sheet_out / "named_ranges.json"

        column_headers = [f"c{column}" for column in range(1, used_cols + 1)]
        with values_path.open("w", encoding="utf-8", newline="") as values_handle:
            values_writer = csv.writer(values_handle)
            values_writer.writerow(["row_number", *column_headers])
            if used_rows > 0 and used_cols > 0:
                for row_idx, row in enumerate(
                    sheet.iter_rows(
                        min_row=1,
                        max_row=used_rows,
                        min_col=1,
                        max_col=used_cols,
                        values_only=False,
                    ),
                    start=1,
                ):
                    values_writer.writerow([row_idx, *[to_text(cell.value) for cell in row]])

        formula_rows: list[list[str | int]] = []
        if used_rows > 0 and used_cols > 0:
            for row_idx, row in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=used_rows,
                    min_col=1,
                    max_col=used_cols,
                    values_only=False,
                ),
                start=1,
            ):
                for col_idx, cell in enumerate(row, start=1):
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_total += 1
                        external = is_external_formula(value)
                        if external:
                            external_formula_total += 1
                        formula_rows.append(
                            [
                                sheet.title,
                                cell.coordinate,
                                row_idx,
                                col_idx,
                                value,
                                "true" if external else "false",
                            ]
                        )

        with formulas_path.open("w", encoding="utf-8", newline="") as formula_handle:
            writer = csv.writer(formula_handle)
            writer.writerow(["sheet", "cell", "row", "col", "formula", "is_external"])
            writer.writerows(formula_rows)

        write_json(named_ranges_path, named_ranges_by_sheet.get(sheet.title, []))

        workbook_meta["sheets"].append(
            {
                "sheet_name": sheet.title,
                "sheet_slug": sheet_slug,
                "max_row": used_rows,
                "max_col": used_cols,
                "formula_cells": len(formula_rows),
                "external_formula_cells": sum(1 for row in formula_rows if row[-1] == "true"),
                "values_csv": str(values_path.relative_to(output_dir).as_posix()),
                "formula_cells_csv": str(formulas_path.relative_to(output_dir).as_posix()),
                "named_ranges_json": str(named_ranges_path.relative_to(output_dir).as_posix()),
            }
        )

    workbook.close()

    lineage_flags = {
        "has_external_links": len(external_links) > 0,
        "external_link_count": len(external_links),
        "formula_cells_total": formula_total,
        "external_formula_cells_total": external_formula_total,
    }

    write_json(output_dir / "workbook_meta.json", workbook_meta)
    write_json(output_dir / "external_links.json", external_links)
    write_json(output_dir / "lineage_flags.json", lineage_flags)

    return {
        "workbook_meta": workbook_meta,
        "external_links": external_links,
        "lineage_flags": lineage_flags,
    }

