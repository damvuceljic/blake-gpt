from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from finance_copilot.workbook import extract_workbook


def test_extract_workbook_basic(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Bridge"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "Variance"
    ws["B2"] = "=1+2"
    ws["A3"] = "External"
    ws["B3"] = "='[ExternalBook.xlsx]Sheet1'!$A$1"
    wb.save(workbook_path)

    output_dir = tmp_path / "extract"
    result = extract_workbook(input_path=workbook_path, output_dir=output_dir)

    assert result["workbook_meta"]["sheet_count"] == 1
    assert (output_dir / "workbook_meta.json").exists()
    assert (output_dir / "lineage_flags.json").exists()
    flags = result["lineage_flags"]
    assert flags["formula_cells_total"] >= 2
    assert flags["external_formula_cells_total"] >= 1

